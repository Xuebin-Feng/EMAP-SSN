import os
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

import numpy as np
import openpyxl
from Bio.Align import MultipleSeqAlignment
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DIR = os.path.join(PROJECT_ROOT, "src")
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)

import Alignment_Manager
from commands import label
from utilities import Application_Windows as application_windows


class AlignmentStub:
    def __init__(self, sequences=None):
        sequences = ["A", "A", "C"] if sequences is None else list(sequences)
        self.aln = MultipleSeqAlignment(
            [
                SeqRecord(Seq(sequence), id=f"node{index}")
                for index, sequence in enumerate(sequences)
            ]
        )
        self.col_to_label = {0: "1"}
        self.label_to_col = {"1": 0}
        self.has_reference = True
        self.resolved_ref_full = "node0"
        self.viewer_to_aln = np.arange(len(sequences), dtype=int)

    def calculate_frequencies(self, mapping, exclude=None, aln=None):
        return Alignment_Manager.calculate_frequencies(
            aln if aln is not None else self.aln,
            mapping,
            exclude or [],
        )


class ImmediateScheduler:
    def __init__(self, viewer):
        self.viewer = viewer

    def is_output_path_reserved(self, _path):
        return False

    def enqueue(self, **job):
        result = job["worker"](job["payload"])
        self.viewer.console_text.text = result["message"]
        reveal_directory = result.get("reveal_directory")
        if reveal_directory:
            application_windows.open_in_file_manager(reveal_directory)
        return 1


class CapturingScheduler:
    def __init__(self):
        self.job = None

    def is_output_path_reserved(self, _path):
        return False

    def enqueue(self, **job):
        self.job = job
        return 1


def find_row(worksheet, first_cell_value):
    for row_idx in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_idx, column=1).value == first_cell_value:
            return row_idx
    raise AssertionError(f"Could not find row {first_cell_value!r}")


class LabelWorkbookPercentTests(unittest.TestCase):
    @staticmethod
    def make_viewer(sequences, cluster_labels, group_labels):
        headers = [f"node{index}" for index in range(len(sequences))]
        return SimpleNamespace(
            alignment=AlignmentStub(sequences),
            active_reference="node0",
            full_headers=headers,
            n_nodes=len(headers),
            cluster_labels=(
                None
                if cluster_labels is None
                else np.asarray(cluster_labels, dtype=int)
            ),
            group_labels=[set(groups) for groups in group_labels],
            console_text=SimpleNamespace(text=""),
            last_cluster_params=None,
        )

    def run_label(self, directory, args, viewer=None, identity_weights=None):
        if viewer is None:
            viewer = SimpleNamespace(
                alignment=AlignmentStub(),
                active_reference="node0",
                full_headers=["node0", "node1", "node2", "unaligned"],
                n_nodes=4,
                cluster_labels=np.array([0, 0, 1, -1]),
                group_labels=[{"GroupA"}, set(), set(), set()],
                console_text=SimpleNamespace(text=""),
                last_cluster_params=None,
            )
        viewer.background_job_scheduler = ImmediateScheduler(viewer)
        metadata = SimpleNamespace(model_name="test-model", network_type="cosine")
        viewer_to_aln = getattr(viewer.alignment, "viewer_to_aln", None)
        if viewer_to_aln is None or np.asarray(viewer_to_aln).shape != (
            len(viewer.full_headers),
        ):
            viewer_to_aln = np.full(len(viewer.full_headers), -1, dtype=int)
            aligned_count = min(len(viewer.alignment.aln), len(viewer.full_headers))
            viewer_to_aln[:aligned_count] = np.arange(aligned_count)
        viewer_to_aln = np.asarray(viewer_to_aln, dtype=int)
        if identity_weights is None:
            identity_weights = np.array([0.5, 0.5, 1.0])

        with mock.patch.object(label.cfg, "CLUSTER_LABEL_DIR", directory), \
                mock.patch.object(label.cfg, "NODE_FASTA_FILE", "nodes.fasta"), \
                mock.patch.object(label.cfg, "INPUT_HDF5", "network.h5"), \
                mock.patch.object(label.cfg, "MSA_FILE", "alignment.fasta"), \
                mock.patch.object(label.cache_manifest, "validate_network_schema", return_value=metadata), \
                mock.patch.object(
                    label.Command_Engine,
                    "get_alignment_mapping",
                    return_value=(
                        viewer_to_aln,
                        np.flatnonzero(viewer_to_aln >= 0),
                    ),
                ), \
                mock.patch.object(label.cluster_cmd, "get_cluster_color_map", return_value={0: (1.0, 0.0, 0.0), 1: (0.0, 1.0, 0.0)}), \
                mock.patch.object(
                    label.logo_cmd,
                    "calculate_identity_weights",
                    return_value=np.asarray(identity_weights, dtype=float),
                ) as identity_weight_mock, \
                mock.patch.object(application_windows, "open_in_file_manager"):
            label.run(viewer, args)
        viewer.identity_weight_mock = identity_weight_mock
        return viewer

    def test_target_aliases_and_default_have_distinct_modes(self):
        self.assertEqual(label._parse_label_arguments([])["forced_target"], "all")
        for target in ("cluster", "clusters"):
            with self.subTest(target=target):
                self.assertEqual(
                    label._parse_label_arguments([target])["forced_target"],
                    "clusters",
                )
        for target in ("group", "groups"):
            with self.subTest(target=target):
                self.assertEqual(
                    label._parse_label_arguments([target])["forced_target"],
                    "groups",
                )

    def test_percent_column_uses_total_network_nodes_in_both_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(directory, [])

            output_path = next(Path(directory).glob("Label_Output_*.xlsx"))
            workbook = openpyxl.load_workbook(output_path)

            for sheet_name in ("Subset Stats", "Occupancy Stats"):
                worksheet = workbook[sheet_name]
                header_row = find_row(worksheet, "Subset Name")
                cluster_row = find_row(worksheet, "Cluster 0")
                group_row = find_row(worksheet, "Group GroupA")
                global_row = find_row(worksheet, "Global Stats")

                self.assertEqual(worksheet.freeze_panes, "C1")
                self.assertEqual(worksheet.cell(header_row, 2).value, "Percent")
                self.assertEqual(worksheet.cell(header_row, 3).value, "Count")
                self.assertEqual(worksheet.cell(header_row, 4).value, "Hex Color")
                self.assertEqual(worksheet.cell(header_row, 10).value, "#1")
                self.assertEqual(worksheet.cell(cluster_row, 2).value, 0.5)
                self.assertEqual(worksheet.cell(group_row, 2).value, 0.25)
                self.assertEqual(worksheet.cell(global_row, 2).value, 0.75)
                self.assertEqual(worksheet.cell(cluster_row, 2).number_format, "0.00%")
                self.assertEqual(worksheet.column_dimensions["B"].width, 10.0)
                self.assertEqual(worksheet.cell(cluster_row, 4).fill.fill_type, "solid")
                self.assertEqual(worksheet.cell(cluster_row, 10).fill.fill_type, "solid")
            viewer = self.run_label(directory, ["second_report"])
            viewer.identity_weight_mock.assert_not_called()

    def test_custom_filename_is_used_and_xlsx_is_appended(self):
        with tempfile.TemporaryDirectory() as directory:
            viewer = self.run_label(
                directory,
                ["0.4", "0.9", "custom_label_report"],
            )

            output_path = Path(directory, "custom_label_report.xlsx")
            self.assertTrue(output_path.is_file())
            self.assertIn(str(output_path), viewer.console_text.text)
            self.assertEqual(
                [path for path in Path(directory).iterdir() if ".partial" in path.name],
                [],
            )

            workbook = openpyxl.load_workbook(output_path)
            metadata_sheet = workbook["Meta Data"]
            alignment_row = find_row(metadata_sheet, "Alignment Name")
            label_row = find_row(metadata_sheet, "Label Parameters")
            statistics_row = find_row(metadata_sheet, "Statistics")
            self.assertEqual(label_row, alignment_row + 1)
            self.assertEqual(statistics_row, label_row + 1)
            self.assertEqual(
                metadata_sheet.cell(label_row, 2).value,
                "gmax_outside=40%, cmin=90%, target=all",
            )
            self.assertEqual(
                metadata_sheet.cell(statistics_row, 2).value,
                "Aligned 3 of 4 | Excluded 1 | Effective 3",
            )
            for sheet_name in ("Subset Stats", "Occupancy Stats"):
                worksheet = workbook[sheet_name]
                self.assertIsNotNone(find_row(
                    worksheet,
                    "Statistics: Aligned 3 of 4 | Excluded 1 | Effective 3",
                ))
                self.assertIsNotNone(find_row(worksheet, "Global Conserved (>97%)"))

    def test_custom_filename_overwrites_existing_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory, "overwrite_report.xlsx")
            output_path.write_bytes(b"previous content")

            viewer = self.run_label(directory, ["overwrite_report"])

            self.assertIn(str(output_path), viewer.console_text.text)
            self.assertNotEqual(output_path.read_bytes(), b"previous content")
            workbook = openpyxl.load_workbook(output_path)
            self.assertIn("Meta Data", workbook.sheetnames)
            self.assertEqual(
                [path for path in Path(directory).iterdir() if ".partial" in path.name],
                [],
            )

    def test_automatic_output_created_during_render_is_preserved(self):
        with tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory, "automatic.xlsx")

            def save_and_create_competing_output(
                _workbook,
                partial_path,
            ):
                Path(partial_path).write_bytes(b"rendered content")
                output_path.write_bytes(b"competing content")

            with mock.patch.object(
                label,
                "_available_automatic_output",
                return_value=(output_path.name, str(output_path)),
            ), mock.patch(
                "openpyxl.workbook.workbook.Workbook.save",
                autospec=True,
                side_effect=save_and_create_competing_output,
            ):
                viewer = self.run_label(directory, [])

            self.assertIn("already exists", viewer.console_text.text)
            self.assertEqual(output_path.read_bytes(), b"competing content")
            self.assertEqual(
                [path for path in Path(directory).iterdir() if ".partial" in path.name],
                [],
            )

    def test_bare_filename_is_allowed_after_keyword_parameters(self):
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(directory, ["cmin", "90%", "keyword_report"])

            self.assertTrue(Path(directory, "keyword_report.xlsx").is_file())

    def test_identity_keyword_forms_are_equivalent(self):
        for index, identity_value in enumerate(("0.9", "90", "90%")):
            with self.subTest(identity_value=identity_value), \
                    tempfile.TemporaryDirectory() as directory:
                viewer = self.run_label(
                    directory,
                    ["id", identity_value, f"identity_{index}"],
                )

                self.assertTrue(Path(directory, f"identity_{index}.xlsx").is_file())
                call_args = viewer.identity_weight_mock.call_args
                self.assertEqual(call_args.args[1], 0.9)
                self.assertTrue(call_args.kwargs["report_backend"])

    def test_third_positional_number_enables_identity_weighting(self):
        with tempfile.TemporaryDirectory() as directory:
            viewer = self.run_label(
                directory,
                ["0.4", "0.9", "100%", "weighted_report"],
            )

            output_path = Path(directory, "weighted_report.xlsx")
            workbook = openpyxl.load_workbook(output_path)
            metadata_rows = [
                (row[0].value, row[1].value)
                for row in workbook["Meta Data"].iter_rows(min_col=1, max_col=2)
            ]
            metadata = dict(metadata_rows)

            self.assertEqual(
                metadata["Statistics"],
                "Aligned 3 of 4 | Excluded 1 | Effective 2",
            )
            self.assertNotIn("Network Nodes", metadata)
            self.assertNotIn("Aligned Nodes", metadata)
            self.assertNotIn("Global Effective N", metadata)
            self.assertNotIn("Excluded Unaligned Nodes", metadata)
            self.assertNotIn("Identity Threshold", metadata)
            self.assertNotIn("Identity Backend", metadata)
            self.assertNotIn("Identity Threads", metadata)
            self.assertNotIn("Identity Fallback Reason", metadata)
            metadata_names = [name for name, _value in metadata_rows]
            alignment_index = metadata_names.index("Alignment Name")
            self.assertEqual(metadata_names[alignment_index + 1], "Label Parameters")
            self.assertEqual(metadata_names[alignment_index + 2], "Statistics")
            self.assertIn("identity=100%", metadata["Label Parameters"])
            self.assertNotIn("outside_pool", metadata["Label Parameters"])
            self.assertNotIn(
                "global_conservation_threshold",
                metadata["Label Parameters"],
            )
            viewer.identity_weight_mock.assert_called_once()

            for sheet_name in ("Subset Stats", "Occupancy Stats"):
                worksheet = workbook[sheet_name]
                self.assertIsNotNone(find_row(
                    worksheet,
                    "Statistics: Aligned 3 of 4 | Excluded 1 | Effective 2",
                ))
                header_row = find_row(worksheet, "Subset Name")
                global_row = find_row(worksheet, "Global Stats")
                cluster_row = find_row(worksheet, "Cluster 0")
                group_row = find_row(worksheet, "Group GroupA")

                self.assertEqual(worksheet.freeze_panes, "C1")
                self.assertEqual(
                    [worksheet.cell(header_row, column).value for column in range(1, 6)],
                    [
                        "Subset Name", "Effective Percent", "Percent",
                        "Effective N", "Count N",
                    ],
                )
                self.assertEqual(worksheet.cell(header_row, 6).value, "Hex Color")
                self.assertEqual(worksheet.cell(header_row, 12).value, "#1")
                self.assertEqual(worksheet.cell(global_row, 2).value, 1.0)
                self.assertEqual(worksheet.cell(global_row, 3).value, 0.75)
                self.assertEqual(worksheet.cell(global_row, 4).value, 2.0)
                self.assertEqual(worksheet.cell(global_row, 5).value, 3)
                self.assertEqual(worksheet.cell(cluster_row, 2).value, 0.5)
                self.assertEqual(worksheet.cell(cluster_row, 3).value, 0.5)
                self.assertEqual(worksheet.cell(cluster_row, 4).value, 1.0)
                self.assertEqual(worksheet.cell(cluster_row, 5).value, 2)
                self.assertEqual(worksheet.cell(group_row, 2).value, 0.25)
                self.assertEqual(worksheet.cell(group_row, 3).value, 0.25)
                self.assertEqual(
                    worksheet.cell(cluster_row, 2).number_format,
                    "0.00%",
                )
                self.assertEqual(
                    worksheet.cell(cluster_row, 3).number_format,
                    "0.00%",
                )
                self.assertEqual(
                    worksheet.cell(cluster_row, 4).number_format,
                    "0.00",
                )
                self.assertEqual(
                    worksheet.cell(cluster_row, 6).fill.fill_type,
                    "solid",
                )

    def test_identity_assignment_errors_do_not_write_workbooks(self):
        cases = (
            (["id"], "Missing numerical value for 'id'"),
            (["id", "0"], "outside the supported range"),
            (["id", "90%", "id", "80%"], "Duplicate assignment for 'id'"),
            (
                ["0.4", "0.9", "90%", "id", "80%"],
                "defined both positionally and via keyword",
            ),
            (["0.4", "0.9", "90%", "80%"], "Too many positional"),
        )
        for args, expected_error in cases:
            with self.subTest(args=args), tempfile.TemporaryDirectory() as directory:
                viewer = self.run_label(directory, args)

                self.assertIn(expected_error, viewer.console_text.text)
                self.assertEqual(list(Path(directory).glob("*.xlsx")), [])
                viewer.identity_weight_mock.assert_not_called()

    def test_numeric_filename_requires_and_accepts_xlsx_extension(self):
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(directory, ["0.4.xlsx"])

            self.assertTrue(Path(directory, "0.4.xlsx").is_file())

    def test_filename_must_be_final_and_old_keyword_form_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            viewer = self.run_label(directory, ["report", "cmin", "90%"])
            self.assertEqual(
                viewer.console_text.text,
                "Error: A custom output filename must be the final argument.",
            )
            self.assertEqual(list(Path(directory).glob("*.xlsx")), [])

        with tempfile.TemporaryDirectory() as directory:
            viewer = self.run_label(directory, ["filename", "old_style"])
            self.assertEqual(
                viewer.console_text.text,
                "Error: A custom output filename must be the final argument.",
            )
            self.assertEqual(list(Path(directory).glob("*.xlsx")), [])

    def test_gmin_is_not_user_configurable(self):
        with tempfile.TemporaryDirectory() as directory:
            viewer = self.run_label(directory, ["gmin", "90%"])

            self.assertEqual(
                viewer.console_text.text,
                "Error: gmin is fixed at 97% and cannot be set by the label command.",
            )
            self.assertEqual(list(Path(directory).glob("*.xlsx")), [])

    def test_shared_union_reports_two_clusters_that_fail_individually(self):
        viewer = self.make_viewer(
            ["Y", "Y", "Y", "Y", "A", "A"],
            [0, 0, 1, 1, 2, 2],
            [set() for _ in range(6)],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["gmax", "50%", "cmin", "100%", "shared_clusters"],
                viewer=viewer,
            )
            workbook = openpyxl.load_workbook(
                Path(directory, "shared_clusters.xlsx")
            )
            worksheet = workbook["Subset Stats"]
            header_row = find_row(worksheet, "Subset Name")
            self.assertEqual(worksheet.cell(header_row, 10).value, "#1")
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Cluster 0"), 10).value,
                "Y1",
            )
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Cluster 1"), 10).value,
                "Y1",
            )

    def test_default_mode_pools_clusters_and_groups_together(self):
        viewer = self.make_viewer(
            ["Y", "Y", "Y", "Y", "A"],
            [0, 0, 1, 1, 1],
            [set(), set(), {"SharedY"}, {"SharedY"}, set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["gmax", "50%", "cmin", "100%", "cross_type"],
                viewer=viewer,
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "cross_type.xlsx")
            )["Subset Stats"]
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Cluster 0"), 10).value,
                "Y1",
            )
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group SharedY"), 10).value,
                "Y1",
            )

    def test_cluster_aliases_exclude_group_results(self):
        for target in ("cluster", "clusters"):
            with self.subTest(target=target), tempfile.TemporaryDirectory() as directory:
                self.run_label(directory, [target, f"{target}_only"])
                worksheet = openpyxl.load_workbook(
                    Path(directory, f"{target}_only.xlsx")
                )["Subset Stats"]
                subset_names = {
                    worksheet.cell(row=row, column=1).value
                    for row in range(1, worksheet.max_row + 1)
                }

                self.assertIn("Cluster 0", subset_names)
                self.assertIn("Cluster 1", subset_names)
                self.assertNotIn("Group GroupA", subset_names)

    def test_group_aliases_exclude_cluster_results(self):
        for target in ("group", "groups"):
            with self.subTest(target=target), tempfile.TemporaryDirectory() as directory:
                self.run_label(directory, [target, f"{target}_only"])
                worksheet = openpyxl.load_workbook(
                    Path(directory, f"{target}_only.xlsx")
                )["Subset Stats"]
                subset_names = {
                    worksheet.cell(row=row, column=1).value
                    for row in range(1, worksheet.max_row + 1)
                }

                self.assertIn("Group GroupA", subset_names)
                self.assertNotIn("Cluster 0", subset_names)
                self.assertNotIn("Cluster 1", subset_names)

    def test_default_mode_uses_groups_when_clusters_are_unavailable(self):
        viewer = self.make_viewer(
            ["Y", "Y", "A"],
            None,
            [{"OnlyGroup"}, {"OnlyGroup"}, set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(directory, ["groups_available"], viewer=viewer)
            worksheet = openpyxl.load_workbook(
                Path(directory, "groups_available.xlsx")
            )["Subset Stats"]

            self.assertIsNotNone(find_row(worksheet, "Group OnlyGroup"))

    def test_group_mode_deduplicates_overlapping_weighted_memberships(self):
        viewer = self.make_viewer(
            ["Y", "Y", "A"],
            None,
            [{"G1"}, {"G1", "G2"}, set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                [
                    "groups", "gmax", "10%", "cmin", "100%",
                    "id", "100%", "overlap",
                ],
                viewer=viewer,
                identity_weights=[0.5, 0.25, 1.0],
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "overlap.xlsx")
            )["Subset Stats"]
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group G1"), 12).value,
                "Y1",
            )
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group G2"), 12).value,
                "Y1",
            )

    def test_different_amino_acids_form_independent_exclusion_pools(self):
        viewer = self.make_viewer(
            ["Y", "Y", "F", "F", "Y", "A"],
            None,
            [{"GY"}, {"GY"}, {"GF"}, {"GF"}, set(), set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["groups", "gmax", "40%", "cmin", "100%", "independent"],
                viewer=viewer,
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "independent.xlsx")
            )["Subset Stats"]
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group GY"), 10).value,
                "Y1",
            )
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group GF"), 10).value,
                "F1",
            )

    def test_subset_below_cmin_is_not_excluded_or_reported(self):
        viewer = self.make_viewer(
            ["Y", "Y", "Y", "A", "A", "A"],
            None,
            [{"High"}, {"High"}, {"Low"}, {"Low"}, set(), set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["groups", "gmax", "20%", "cmin", "75%", "below_cmin"],
                viewer=viewer,
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "below_cmin.xlsx")
            )["Subset Stats"]
            header_row = find_row(worksheet, "Subset Name")
            self.assertIsNone(worksheet.cell(header_row, 10).value)

    def test_empty_shared_background_is_not_reported(self):
        viewer = self.make_viewer(
            ["Y", "Y"],
            None,
            [{"G1"}, {"G2"}],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["groups", "cmin", "100%", "empty_background"],
                viewer=viewer,
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "empty_background.xlsx")
            )["Subset Stats"]
            header_row = find_row(worksheet, "Subset Name")
            self.assertIsNone(worksheet.cell(header_row, 10).value)

    def test_multiple_passing_amino_acids_share_one_ordered_cell(self):
        viewer = self.make_viewer(
            ["Y", "Y", "F", "A", "A"],
            None,
            [{"Mixed"}, {"Mixed"}, {"Mixed"}, set(), set()],
        )
        with tempfile.TemporaryDirectory() as directory:
            self.run_label(
                directory,
                ["groups", "gmax", "40%", "cmin", "30%", "multi_residue"],
                viewer=viewer,
            )
            worksheet = openpyxl.load_workbook(
                Path(directory, "multi_residue.xlsx")
            )["Subset Stats"]
            self.assertEqual(
                worksheet.cell(find_row(worksheet, "Group Mixed"), 10).value,
                "Y1 | F1",
            )

    def test_filename_cannot_escape_output_directory(self):
        with self.assertRaisesRegex(ValueError, "path separators"):
            label._normalize_output_filename("../outside")

    def test_automatic_filename_uses_deterministic_numeric_suffix(self):
        with tempfile.TemporaryDirectory() as directory:
            first = os.path.join(directory, "Label_Output_stamp.xlsx")
            second = os.path.join(directory, "Label_Output_stamp_2.xlsx")
            Path(first).write_bytes(b"existing")
            scheduler = CapturingScheduler()
            scheduler.is_output_path_reserved = lambda path: (
                os.path.normcase(os.path.abspath(path))
                == os.path.normcase(os.path.abspath(second))
            )

            filename, path = label._available_automatic_output(
                scheduler,
                directory,
                "Label_Output_stamp.xlsx",
            )

            self.assertEqual(filename, "Label_Output_stamp_3.xlsx")
            self.assertEqual(path, os.path.join(directory, filename))

    def test_enqueued_label_keeps_invocation_time_alignment_and_memberships(self):
        with tempfile.TemporaryDirectory() as directory:
            alignment = AlignmentStub()
            alignment.viewer_to_aln = np.array([0, 1, 2, -1])
            scheduler = CapturingScheduler()
            viewer = SimpleNamespace(
                alignment=alignment,
                active_reference="node0",
                alignment_offset=7,
                full_headers=["node0", "node1", "node2", "unaligned"],
                n_nodes=4,
                cluster_labels=np.array([0, 0, 1, -1]),
                group_labels=[{"GroupA"}, set(), set(), set()],
                console_text=SimpleNamespace(text=""),
                last_cluster_params=("leiden_1.0", 10),
                background_job_scheduler=scheduler,
            )
            metadata = SimpleNamespace(
                model_name="test-model",
                network_type="cosine",
            )
            with mock.patch.object(label.cfg, "CLUSTER_LABEL_DIR", directory), \
                    mock.patch.object(label.cfg, "INPUT_HDF5", "network.h5"), \
                    mock.patch.object(
                        label.cache_manifest,
                        "validate_network_schema",
                        return_value=metadata,
                    ):
                label.run(viewer, ["snapshot_report"])
                explicit_job = scheduler.job
                label.run(viewer, [])
                automatic_job = scheduler.job

            alignment.aln[0].seq = Seq("G")
            alignment.col_to_label[0] = "99"
            viewer.cluster_labels[0] = 9
            viewer.group_labels[0].add("LaterGroup")
            viewer.active_reference = "node2"
            viewer.alignment_offset = 99

            snapshot = explicit_job["payload"].viewer_snapshot
            self.assertEqual(str(snapshot.alignment.aln[0].seq), "A")
            self.assertEqual(snapshot.alignment.col_to_label, {0: "1"})
            self.assertEqual(snapshot.cluster_labels, (0, 0, 1, -1))
            self.assertEqual(snapshot.group_labels[0], frozenset({"GroupA"}))
            self.assertEqual(snapshot.active_reference, "node0")
            self.assertEqual(snapshot.alignment_offset, 7)
            self.assertTrue(explicit_job["allow_overwrite"])
            self.assertTrue(snapshot._label_allow_overwrite)
            self.assertFalse(automatic_job["allow_overwrite"])
            self.assertFalse(
                automatic_job["payload"].viewer_snapshot._label_allow_overwrite
            )


if __name__ == "__main__":
    unittest.main()
