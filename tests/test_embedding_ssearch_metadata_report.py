"""Regression coverage for metadata-viewer-compatible SSEARCH workbooks."""

import os
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from io import StringIO
from types import SimpleNamespace
from unittest import mock

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DIR = os.path.join(PROJECT_ROOT, "src")
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)

import Command_Engine
from tools import Embedding_SSEARCH
from web_ui import meta_backend


class EmbeddingSsearchMetadataReportTests(unittest.TestCase):
    def test_xlsx_matches_metadata_template_and_imports_into_viewer(self):
        results = pd.DataFrame(
            [
                {
                    "index": -1,
                    "header": "(Query) query_node",
                    "raw_score": 0.0,
                    "norm_score": 99.9,
                    "length": 90,
                    "seq_len": 90,
                    "aln_len": 90,
                },
                {
                    "index": 4,
                    "header": "node_alpha",
                    "raw_score": 42.5,
                    "norm_score": 0.625,
                    "length": 100,
                    "seq_len": 100,
                    "aln_len": 85,
                },
                {
                    "index": 7,
                    "header": "node_beta",
                    "raw_score": 37.25,
                    "norm_score": 0.5,
                    "length": 120,
                    "seq_len": 120,
                    "aln_len": 90,
                },
            ]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            with (
                mock.patch.object(Embedding_SSEARCH, "REPORT_DIR", temp_dir),
                mock.patch.object(Embedding_SSEARCH, "INPUT_EMBED", "db.h5"),
                mock.patch.object(Embedding_SSEARCH, "TOP_K", 10),
                mock.patch.object(Embedding_SSEARCH, "NORM_THRESHOLD", None),
                mock.patch.object(Embedding_SSEARCH, "ALIGNMENT_MODE", "local"),
                mock.patch.object(Embedding_SSEARCH, "GENERATE_FASTA", False),
                redirect_stdout(StringIO()),
            ):
                Embedding_SSEARCH.save_results(
                    results,
                    ("query_node", 90),
                    3,
                    {},
                    "metadata_test",
                    "A" * 90,
                    "alignment_length",
                    -2.0,
                )

            output_path = os.path.join(temp_dir, "Report_metadata_test.xlsx")
            self.assertTrue(os.path.isfile(output_path))

            raw = pd.read_excel(
                output_path,
                sheet_name="Search Results",
                header=None,
            )
            self.assertEqual(
                raw.iloc[0].tolist(),
                [
                    "Node ID",
                    "Rank",
                    "Norm_Score",
                    "Raw_Score",
                    "Sequence_Length",
                    "Alignment_Length",
                ],
            )
            self.assertEqual(
                raw.iloc[1].tolist(),
                ["Data Type", "number", "number", "number", "number", "number"],
            )
            self.assertEqual(raw.iloc[2, 0], "node_alpha")
            self.assertEqual(raw.iloc[3, 0], "node_beta")
            self.assertNotIn("(Query) query_node", raw.iloc[:, 0].tolist())

            workbook = load_workbook(output_path, read_only=False)
            self.assertEqual(workbook.sheetnames, ["Search Results", "Search Parameters"])
            worksheet = workbook["Search Results"]
            self.assertEqual(worksheet["A1"].fill.fgColor.rgb, "002C3E50")
            self.assertEqual(worksheet["A2"].fill.fgColor.rgb, "00D5D8DC")
            self.assertEqual(worksheet.freeze_panes, "A3")
            self.assertEqual(worksheet["C3"].number_format, "0.000")
            parameters = workbook["Search Parameters"]
            self.assertEqual(parameters["A1"].fill.fgColor.rgb, "002C3E50")
            self.assertEqual(parameters.column_dimensions["B"].width, 80)
            workbook.close()

            viewer = SimpleNamespace(
                full_headers=["node_alpha", "node_beta"],
                n_nodes=2,
                metadata={},
                visible_mask=np.ones(2, dtype=bool),
                selected_indices=[],
                _save_state=mock.Mock(),
                broadcast_event=mock.Mock(),
                get_serializable_metadata=lambda: {},
            )
            with mock.patch.object(Command_Engine, "print_help") as print_help:
                meta_backend.upload_metadata(viewer, [output_path])

            self.assertEqual(
                list(viewer.metadata),
                [
                    "Rank",
                    "Norm_Score",
                    "Raw_Score",
                    "Sequence_Length",
                    "Alignment_Length",
                ],
            )
            self.assertTrue(
                all(entry["type"] == "number" for entry in viewer.metadata.values())
            )
            np.testing.assert_allclose(viewer.metadata["Rank"]["values"], [1, 2])
            np.testing.assert_allclose(
                viewer.metadata["Norm_Score"]["values"],
                [0.625, 0.5],
            )
            print_help.assert_called_once()


if __name__ == "__main__":
    unittest.main()
