import importlib
import os
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

import numpy as np
import openpyxl
from Bio.Align import MultipleSeqAlignment
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_DIR = os.path.join(PROJECT_ROOT, "src")
if SRC_DIR not in sys.path:
    sys.path.insert(0, SRC_DIR)

import Alignment_Manager
from commands import specificity


class AlignmentStub:
    def __init__(self, sequences):
        self.aln = MultipleSeqAlignment(
            [SeqRecord(Seq(sequence), id=f"node{index}") for index, sequence in enumerate(sequences)]
        )
        self.col_to_label = {index: str(index + 1) for index in range(len(sequences[0]))}
        self.label_to_col = {label: column for column, label in self.col_to_label.items()}
        self.has_reference = True
        self.resolved_ref_full = "node0"

    def calculate_frequencies(self, mapping, exclude=None, aln=None):
        return Alignment_Manager.calculate_frequencies(
            aln if aln is not None else self.aln,
            mapping,
            exclude or [],
        )


class SpecificityParserTests(unittest.TestCase):
    def test_defaults_and_keyword_values(self):
        defaults = specificity.parse_command_args([])
        self.assertEqual(defaults.target, "clusters")
        self.assertEqual(defaults.delta, 0.15)
        self.assertEqual(defaults.qmax, 0.05)
        self.assertEqual(defaults.posterior, 0.95)
        self.assertIsNone(defaults.identity)
        self.assertIsNone(defaults.filename)

        parsed = specificity.parse_command_args(
            ["groups", "delta", "10%", "qmax", "0.1", "posterior", "90%", "identity", "80%", "report"]
        )
        self.assertEqual(parsed.target, "groups")
        self.assertEqual(parsed.delta, 0.10)
        self.assertEqual(parsed.qmax, 0.10)
        self.assertEqual(parsed.posterior, 0.90)
        self.assertEqual(parsed.identity, 0.80)
        self.assertEqual(parsed.filename, "report.xlsx")

    def test_identity_uses_logo_parser(self):
        with mock.patch.object(specificity.logo_cmd, "parse_identity_threshold", return_value=0.825) as parser:
            parsed = specificity.parse_command_args(["identity", "82.5%"])
        parser.assert_called_once_with("82.5%")
        self.assertEqual(parsed.identity, 0.825)

    def test_rejects_legacy_duplicate_names_and_invalid_ranges(self):
        invalid = (
            ["cmin", "90%"],
            ["gmin", "90%"],
            ["delta", "10%", "delta", "20%"],
            ["one", "two"],
            ["../report"],
            ["bad:name"],
            ["delta", "101%"],
            ["clusters", "groups"],
            ["posterior"],
        )
        for arguments in invalid:
            with self.subTest(arguments=arguments), self.assertRaises(ValueError):
                specificity.parse_command_args(arguments)

    def test_dynamic_command_module_is_importable(self):
        module = importlib.import_module("commands.specificity")
        self.assertTrue(callable(module.run))


class SpecificityStatisticsTests(unittest.TestCase):
    def test_benjamini_hochberg_preserves_order_and_nan(self):
        adjusted = specificity.benjamini_hochberg([0.01, np.nan, 0.04, 0.03])
        np.testing.assert_allclose(adjusted[[0, 2, 3]], [0.03, 0.04, 0.04])
        self.assertTrue(np.isnan(adjusted[1]))

    def test_jensen_shannon_identical_and_disjoint_profiles(self):
        self.assertAlmostEqual(
            specificity.jensen_shannon_divergence([4, 1], [8, 2]),
            0.0,
        )
        self.assertAlmostEqual(
            specificity.jensen_shannon_divergence([5, 0], [0, 7]),
            1.0,
        )
        self.assertTrue(
            np.isnan(specificity.jensen_shannon_divergence([0, 0], [1, 0]))
        )

    def test_log_odds_is_finite_for_zero_cells(self):
        positive = specificity.smoothed_log2_odds_ratio(10, 10, 0, 10)
        negative = specificity.smoothed_log2_odds_ratio(0, 10, 10, 10)
        self.assertTrue(np.isfinite(positive))
        self.assertAlmostEqual(positive, -negative)
        self.assertGreater(positive, 0.0)

    def test_large_dataset_has_stronger_posterior_evidence_for_same_effect(self):
        small = specificity.beta_difference_probability(4, 5, 1, 10, 0.15)
        large = specificity.beta_difference_probability(40, 50, 10, 100, 0.15)
        self.assertGreater(large, small)
        self.assertGreater(large, 0.99)

    def test_moderate_and_secondary_enrichment_are_flagged_without_conservation_gate(self):
        # Group composition: A=60%, G=40%; outside: A=0%, G=0%, C=100%.
        sequences = ["A"] * 12 + ["G"] * 8 + ["C"] * 80
        encoded = specificity._encode_alignment(sequences)
        tasks = [
            {
                "type": "group",
                "id": "test",
                "name": "Group test",
                "indices": np.arange(20),
                "sort_key": (1, -20, "test"),
            }
        ]
        subsets, columns, residues = specificity._analyze_tasks(
            encoded,
            sequences,
            np.ones(len(sequences)),
            tasks,
            [(0, "1")],
            specificity.SpecificityOptions(identity=None),
        )

        flagged = {row["amino_acid"] for row in residues if row["interesting"]}
        self.assertEqual(flagged, {"A", "G"})
        self.assertIn("A1", subsets[0]["flagged_data"]["1"])
        self.assertIn("G1", subsets[0]["flagged_data"]["1"])
        self.assertEqual(columns[0]["status"], "Analyzed")
        self.assertGreater(columns[0]["js_divergence"], 0.5)

    def test_equal_distributions_are_not_flagged(self):
        sequences = ["A", "A", "C", "C", "A", "A", "C", "C"]
        tasks = [
            {
                "type": "group",
                "id": "same",
                "name": "Group same",
                "indices": np.arange(4),
                "sort_key": (1, -4, "same"),
            }
        ]
        subsets, columns, residues = specificity._analyze_tasks(
            specificity._encode_alignment(sequences),
            sequences,
            np.ones(len(sequences)),
            tasks,
            [(0, "1")],
            specificity.SpecificityOptions(identity=None),
        )
        self.assertFalse(any(row["interesting"] for row in residues))
        self.assertEqual(subsets[0]["flagged_data"], {})
        self.assertAlmostEqual(columns[0]["js_divergence"], 0.0)

    def test_gaps_are_separate_and_ambiguous_symbols_are_excluded_from_composition(self):
        sequences = ["A", "A", "-", "X", "C", "C", "-", "X"]
        tasks = [
            {
                "type": "group",
                "id": "mixed",
                "name": "Group mixed",
                "indices": np.arange(4),
                "sort_key": (1, -4, "mixed"),
            }
        ]
        _subsets, columns, residues = specificity._analyze_tasks(
            specificity._encode_alignment(sequences),
            sequences,
            np.ones(len(sequences)),
            tasks,
            [(0, "1")],
            specificity.SpecificityOptions(identity=None, delta=0.10),
        )
        self.assertAlmostEqual(columns[0]["group_occupancy"], 0.75)
        self.assertAlmostEqual(columns[0]["outside_occupancy"], 0.75)
        self.assertEqual(columns[0]["raw_group_ambiguous"], 1)
        self.assertEqual(columns[0]["raw_outside_ambiguous"], 1)
        self.assertEqual({row["amino_acid"] for row in residues}, {"A", "C"})

    def test_whole_dataset_group_reports_no_outside_and_never_flags(self):
        sequences = ["A", "A", "C"]
        tasks = [
            {
                "type": "group",
                "id": "all",
                "name": "Group all",
                "indices": np.arange(3),
                "sort_key": (1, -3, "all"),
            }
        ]
        _subsets, columns, residues = specificity._analyze_tasks(
            specificity._encode_alignment(sequences),
            sequences,
            np.ones(3),
            tasks,
            [(0, "1")],
            specificity.SpecificityOptions(identity=None),
        )
        self.assertEqual(columns[0]["status"], "No outside sequences")
        self.assertFalse(any(row["interesting"] for row in residues))


class SpecificityWorkbookTests(unittest.TestCase):
    def test_run_weights_once_and_writes_five_sheet_report(self):
        sequences = ["A"] * 20 + ["C"] * 20
        alignment = AlignmentStub(sequences)
        viewer = SimpleNamespace(
            alignment=alignment,
            active_reference="node0",
            full_headers=[f"node{index}" for index in range(40)],
            n_nodes=40,
            cluster_labels=np.array([0] * 20 + [1] * 20),
            group_labels=[set() for _ in range(40)],
            console_text=SimpleNamespace(text=""),
        )
        mapping = np.arange(40, dtype=np.int64)
        metadata = SimpleNamespace(model_name="test-model", network_type="cosine")

        with tempfile.TemporaryDirectory() as directory:
            with mock.patch.object(specificity, "CLUSTER_LABEL_DIRECTORY", directory), \
                    mock.patch.object(specificity.cfg, "NODE_FASTA_FILE", "nodes.fasta"), \
                    mock.patch.object(specificity.cfg, "INPUT_HDF5", "network.h5"), \
                    mock.patch.object(specificity.cfg, "MSA_FILE", "alignment.fasta"), \
                    mock.patch.object(specificity.cache_manifest, "validate_network_schema", return_value=metadata), \
                    mock.patch.object(specificity.Command_Engine, "get_alignment_mapping", return_value=(mapping, mapping)), \
                    mock.patch.object(specificity.cluster_cmd, "get_cluster_color_map", return_value={0: (1.0, 0.0, 0.0), 1: (0.0, 1.0, 0.0)}), \
                    mock.patch.object(specificity.logo_cmd, "calculate_identity_weights", return_value=(np.ones(40), {"backend": "mock", "threads": 1, "fallback_reason": None})) as weight_mock, \
                    mock.patch.object(specificity.utils, "open_in_file_manager"):
                specificity.run(
                    viewer,
                    ["identity", "90%", "custom_specificity"],
                )

            weight_mock.assert_called_once()
            output_paths = list(Path(directory).glob("custom_specificity.xlsx"))
            self.assertEqual(len(output_paths), 1)
            workbook = openpyxl.load_workbook(output_paths[0])
            self.assertEqual(
                workbook.sheetnames,
                ["Meta Data", "Subset Stats", "Occupancy Stats", "Column Details", "Residue Details"],
            )
            self.assertIn("interesting group-enriched residues", viewer.console_text.text)

            metadata_values = {
                workbook["Meta Data"].cell(row, 1).value: workbook["Meta Data"].cell(row, 2).value
                for row in range(1, workbook["Meta Data"].max_row + 1)
            }
            self.assertEqual(metadata_values["Command"], "specificity")
            self.assertEqual(metadata_values["Identity Threshold"], 0.9)
            self.assertEqual(metadata_values["Identity Backend"], "mock")
            self.assertEqual(metadata_values["Global Conservation Threshold"], 0.97)
            self.assertIn("not independent evidence", metadata_values["Inference Warning"])

            subset_values = [
                cell.value
                for row in workbook["Subset Stats"].iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            ]
            self.assertTrue(any(value.startswith("A1 (+100.0 pp)") for value in subset_values))
            self.assertTrue(any(value.startswith("C1 (+100.0 pp)") for value in subset_values))

            residue_sheet = workbook["Residue Details"]
            headers = [cell.value for cell in residue_sheet[1]]
            interesting_column = headers.index("Interesting") + 1
            self.assertTrue(
                any(
                    residue_sheet.cell(row, interesting_column).value is True
                    for row in range(2, residue_sheet.max_row + 1)
                )
            )


if __name__ == "__main__":
    unittest.main()
