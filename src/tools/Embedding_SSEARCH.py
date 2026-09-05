# Copyright 2026 Xuebin Feng
# Author affiliation: University of Toronto
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
File: Embedding_SSEARCH.py
===================================
Similar to the NCBI's traditional SSEARCH program, this tool performs a rigorous one-vs-all database search. However, instead of using a standard amino acid substitution matrix (like BLOSUM62), it aligns sequences based on the high-dimensional structural similarity of their Protein Language Model (pLM) embeddings.

This allows you to find remote homologs that share structural similarities even if their literal sequence identity has degraded entirely.

How it Works:
1. Target Database: You select a complete metadata-first embedding database (.h5) to search against.
2. Query Input: You can either type the header of a stored sequence, OR enable 'Manual Query Seq' and paste a new raw amino acid sequence into the 'Query Sequence' box.
3. Inference: The script calculates the structural embedding for your query.
4. Scanning: Using parallel CPU workers, it scans your query against every sequence in the database using either Local (Smith-Waterman) or Global (Needleman-Wunsch) dynamic programming.
5. Scoring: The raw alignment scores are normalized (to prevent bias toward excessively long sequences) and ranked.

Outputs:
The script generates report files in the configured report directory:
- Report_<Query>.txt: A human-readable text file showing the ranked hits, their normalized scores, raw scores, and effective alignment lengths.
- Hits_<Query>.fasta: A clean FASTA file containing the sequences of all your top hits, ordered strictly by rank, with your query sequence pinned to the very top. This file is perfectly formatted to be immediately dropped into an MSA tool!

Key Parameters:
- Manual Query Seq: When enabled, the script ignores stored-header lookup and sanitizes the supplied Query Sequence in memory.
- Norm Score Cutoff: Filters out any hits that fall below a specific normalized similarity score.
- Alignment Mode: Use 'local' if you are searching for a specific structural domain within larger proteins. Use 'global' if you are comparing overall holistic similarity.
"""
# %%
import os

try:
    from tools import _bootstrap
except ModuleNotFoundError:
    import _bootstrap

import h5py
import numpy as np
import pandas as pd
import torch
import re
import sys
import gc
import threading
import time
from utilities import Hardware_Utils
from utilities.Alignment_Score_Kernels import global_score_length, local_score_length
from utilities.Embedding_Alignment_Engine import (
    BF16_PER_RESIDUE_TOLERANCE,
    EmbeddingTileStore,
    bf16_accelerator_support,
    compute_score_matrix_torch as _shared_score_matrix,
    cuda_matmul_precision,
    cuda_memory_plan,
    estimate_fixed_query_cuda_working_set,
    is_nvidia_cuda,
    normalize_precision_setting,
    precision_element_bytes,
    run_fixed_query_cuda_pipeline,
)
from collections import deque
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from contextlib import nullcontext
from multiprocessing import Pool, set_start_method
from tqdm import tqdm

from utilities.FASTA_Sanitization import sanitize_header, sanitize_sequence
from utilities.Embedding_HDF5 import (
    dtype_for_saving_mode,
    read_embedding_manifest,
    validate_embedding_array,
)
from tools.Generate_Embeddings import find_model_plugin

# ==========================================
# CONFIGURATION
# ==========================================
INPUT_EMBED = "Sample_[E1_RA]_embeddings.h5"

# QUERY SETTINGS
QUERY_HEADER = "Query_Header"
MANUAL_QUERY_SEQ = False
QUERY_SEQUENCE = "" # Used only when MANUAL_QUERY_SEQ is enabled.
OUTPUT_NAME = "" # Optional: Custom base name for the generated output files.

# SEARCH PARAMETERS
TOP_K = 2500 
NORM_THRESHOLD = None
ALIGNMENT_MODE = "local"
LOCAL_GAP_P = -2.0
GLOBAL_GAP_P = 0.0
NORM_MODE = "longer_sequence"

# HARDWARE & CACHE
WORKERS = 8                  
DEVICE_SELECTION = "auto"
ACCELERATOR_PRECISION = "automatic_32bit"
ACCELERATOR_LANES = "auto"
ACCELERATOR_TUNE_PAIRS = 256
TILED_SEARCH_MIN_TARGETS = 512
TF32_SEARCH_MIN_TARGETS = 4096
from utilities.Tool_Directories import project_directory_defaults
from utilities.Tool_Settings import inherited_settings_path, load_tool_settings

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
_DEFAULT_DIRECTORIES = project_directory_defaults(PROJECT_ROOT)
EMBED_DIR = _DEFAULT_DIRECTORIES["EMBED_DIR"]
REPORT_DIR = _DEFAULT_DIRECTORIES["REPORT_DIR"]
GENERATE_FASTA = False

# --- JSON Settings Override ---
import json
import ast

SETTINGS_FILE = inherited_settings_path(__file__) or os.path.join(PROJECT_ROOT, "tools_settings.json")

if __name__ != "__main__" and os.path.exists(SETTINGS_FILE):
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            all_settings = json.load(f)
            
            if "DIRECTORIES" in all_settings:
                for k, v in all_settings["DIRECTORIES"].items():
                    if k in globals() and v is not None and str(v).strip() != "":
                        if not os.path.isabs(str(v)):
                            v = os.path.normpath(os.path.join(PROJECT_ROOT, str(v)))
                        globals()[k] = v
                        
            script_name = os.path.basename(__file__)
            if script_name in all_settings:
                user_settings = all_settings[script_name]
                for k, v in user_settings.items():
                    if k in globals() and v is not None and str(v).strip() != "":
                        orig = globals()[k]
                        if isinstance(orig, int) and not isinstance(orig, bool):
                            try: v = int(v)
                            except: pass
                        elif isinstance(orig, float):
                            try: v = float(v)
                            except: pass
                        elif isinstance(orig, list):
                            try: v = ast.literal_eval(v) if isinstance(v, str) else v
                            except: pass
                        elif orig is None:
                            if v == "None": v = None
                            elif str(v).replace('.', '', 1).isdigit():
                                v = float(v) if '.' in str(v) else int(v)
                        if isinstance(v, str) and k.endswith("_DIR") and not os.path.isabs(v):
                            v = os.path.normpath(os.path.join(PROJECT_ROOT, v))
                        globals()[k] = v
    except Exception as e:
        print(f"Failed to load user settings: {e}")

# --- DYNAMIC INFERENCE ---
FULL_INPUT_EMBED = os.path.join(EMBED_DIR, INPUT_EMBED) if EMBED_DIR else ""

# --- 3. DATA & MODEL LOADING --------------------------------------------------
def load_model_integrated(model_name):
    device = Hardware_Utils.get_optimal_device()
    print(f"[System] Loading model '{model_name}' on {device}...")
    plugin = find_model_plugin(model_name)
    if plugin is None:
        raise ValueError(
            f"Model '{model_name}' is not supported by an available pLM plugin."
        )
    return plugin.load_model(model_name, device), device, plugin

def get_embedding_integrated(seq, model_obj, device, model_type, target_dtype):
    return model_type.get_embedding(seq, model_obj, device, target_dtype)


def resolve_manual_query_sequence(enabled, sequence):
    """Return a canonical manual query only when its switch is enabled."""
    if not enabled:
        return ""
    cleaned_sequence, _, _ = sanitize_sequence(str(sequence or ""))
    if not cleaned_sequence:
        raise ValueError("QUERY_SEQUENCE is empty after sanitization.")
    return cleaned_sequence


def filter_ranked_hits(df, query_name, manual_query_enabled, top_k):
    """Apply self-hit removal and hit limits using the explicit query source."""
    query_from_database = not bool(manual_query_enabled)
    if query_from_database:
        df = df[df['header'] != query_name]

    if top_k is not None:
        # Stored queries occupy one of TOP_K output slots. Manual queries are
        # pinned separately and retain TOP_K database hits, including a
        # same-header record.
        limit = int(top_k) - 1 if query_from_database else int(top_k)
        df = df.head(max(0, limit))

    return df


def release_accelerator_cache(device):
    """Release cached model allocations before the alignment search begins."""
    if device is None:
        return
    device_type = getattr(device, "type", str(device).split(":", 1)[0])
    try:
        if device_type == "cuda":
            torch.cuda.empty_cache()
        elif device_type == "xpu" and hasattr(torch, "xpu"):
            torch.xpu.empty_cache()
        elif device_type == "mps" and hasattr(torch, "mps"):
            torch.mps.empty_cache()
    except (AttributeError, RuntimeError):
        pass


def prepare_database_embeddings():
    if not os.path.exists(FULL_INPUT_EMBED):
        raise FileNotFoundError(
            f"Embedding database not found: {FULL_INPUT_EMBED}. Generate it "
            "with Generate_Embeddings.py before running SSEARCH."
        )
    print(f"[Init] Loading HDF5 embeddings from:\n       {FULL_INPUT_EMBED}")
    with h5py.File(FULL_INPUT_EMBED, "r") as hf:
        return read_embedding_manifest(hf, require_complete=True)

# --- 4. ALIGNMENT & NORMALIZATION LOGIC ---------------------------------------
def compute_score_matrix_torch(emb_i, emb_j, device):
    return _shared_score_matrix(emb_i, emb_j, device, precision="float32")

def normalize_score(raw_score, align_len, len_i, len_j, mode):
    if mode == "alignment_length": return raw_score / align_len if align_len > 0 else 0.0
    elif mode == "shorter_sequence": denom = min(len_i, len_j); return raw_score / denom if denom > 0 else 0.0
    elif mode == "longer_sequence": denom = max(len_i, len_j); return raw_score / denom if denom > 0 else 0.0
    elif mode == "average_sequence": denom = (len_i + len_j) / 2.0; return raw_score / denom if denom > 0 else 0.0
    else: return raw_score / align_len if align_len > 0 else 0.0

# --- HDF5 MULTIPROCESSING INITIALIZATION ---
worker_hf = None
worker_device = None

def init_worker(h5_path):
    global worker_hf, worker_device
    worker_hf = h5py.File(h5_path, "r", libver='latest', swmr=True)
    worker_device = torch.device("cpu")


def finish_search(args):
    idx, header, len_q, len_t, mode, gap, norm_mode, mat = args
    is_local = (mode == "local")
    if is_local:
        raw, path_len = local_score_length(mat, gap)
    else:
        raw, path_len = global_score_length(mat, gap)
    norm = normalize_score(raw, path_len, len_q, len_t, norm_mode)

    if norm_mode == "alignment_length":
        eff_len = path_len
    elif norm_mode == "shorter_sequence":
        eff_len = min(len_q, len_t)
    elif norm_mode == "longer_sequence":
        eff_len = max(len_q, len_t)
    elif norm_mode == "average_sequence":
        eff_len = (len_q + len_t) / 2.0
    else:
        eff_len = path_len

    return {
        "index": idx,
        "header": header,
        "raw_score": raw,
        "norm_score": norm,
        "length": eff_len,
        "seq_len": len_t,
        "aln_len": path_len,
    }


def search_cpu_worker(args):
    idx, header, safe_h, q_emb, mode, gap, norm_mode = args
    global worker_hf, worker_device
    t_emb = worker_hf["embeddings"][safe_h][:]
    mat = compute_score_matrix_torch(q_emb, t_emb, worker_device)
    return finish_search(
        (
            idx,
            header,
            q_emb.shape[0],
            t_emb.shape[0],
            mode,
            gap,
            norm_mode,
            mat,
        )
    )


accelerator_thread_state = threading.local()
accelerator_lane_cache = {}


def _device_type(device):
    return getattr(device, "type", str(device).split(":", 1)[0])


def _uses_accelerator(device):
    return _device_type(device) != "cpu"


def _supports_explicit_streams(device):
    return _device_type(device) in {"cuda", "xpu"}


def _lane_candidates(device, cpu_workers):
    device_type = _device_type(device)
    max_lanes = max(1, int(cpu_workers))
    if device_type == "cuda":
        base, backend_cap = [1, 2, 4, 8, 16], 16
    elif device_type == "xpu":
        base, backend_cap = [1, 2, 4], 4
    else:
        return [1]

    limit = min(max_lanes, backend_cap)
    candidates = [count for count in base if count <= limit]
    if limit not in candidates:
        candidates.append(limit)
    return sorted(set(candidates))


def _configured_lanes(device, cpu_workers):
    configured = ACCELERATOR_LANES
    if isinstance(configured, str):
        normalized = configured.strip().lower()
        if normalized in {"", "auto"}:
            return None
        try:
            configured = int(normalized)
        except ValueError:
            print(
                f"⚠️ Invalid ACCELERATOR_LANES={configured!r}; using auto."
            )
            return None
    try:
        configured = int(configured)
    except (TypeError, ValueError):
        return None
    if configured < 1:
        return None
    if not _supports_explicit_streams(device):
        return 1
    return min(configured, max(1, int(cpu_workers)))


def _accelerator_name(device):
    device_type = _device_type(device)
    try:
        if device_type == "cuda":
            return torch.cuda.get_device_name(device)
        if device_type == "xpu":
            return torch.xpu.get_device_name(device)
        if device_type == "mps" and hasattr(torch.backends.mps, "get_name"):
            return torch.backends.mps.get_name()
    except (AttributeError, RuntimeError):
        pass
    return str(device)


class _DeviceStreamContext:
    def __init__(self, backend, device, stream):
        self.device_context = backend.device(device)
        self.stream_context = backend.stream(stream)

    def __enter__(self):
        self.device_context.__enter__()
        try:
            return self.stream_context.__enter__()
        except BaseException:
            self.device_context.__exit__(*sys.exc_info())
            raise

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            return self.stream_context.__exit__(
                exc_type,
                exc_value,
                traceback,
            )
        finally:
            self.device_context.__exit__(exc_type, exc_value, traceback)


def _stream_context(device):
    device_type = _device_type(device)
    if device_type == "cuda":
        backend = torch.cuda
    elif device_type == "xpu":
        backend = torch.xpu
    else:
        return nullcontext()

    key = (device_type, str(device))
    stream = getattr(accelerator_thread_state, "stream", None)
    if (
        stream is None
        or getattr(accelerator_thread_state, "stream_key", None) != key
    ):
        stream = backend.Stream(device=device)
        accelerator_thread_state.stream = stream
        accelerator_thread_state.stream_key = key
    return _DeviceStreamContext(backend, device, stream)


def _compute_accelerated_search(args):
    idx, header, q_emb, t_emb, mode, gap, norm_mode, device, precision = args
    with torch.inference_mode():
        with _stream_context(device):
            mat = _shared_score_matrix(
                q_emb, t_emb, device, precision=precision
            )
    return (
        idx,
        header,
        q_emb.shape[0],
        t_emb.shape[0],
        mode,
        gap,
        norm_mode,
        mat,
    )


def _run_accelerated_search(
    tasks,
    workers,
    input_h5,
    device,
    lanes,
    show_progress,
    precision="float32",
):
    results = []
    accelerator_pending = set()
    cpu_pending = set()
    ready_for_cpu = deque()
    ready_limit = max(lanes, min(workers, 8))
    task_iterator = iter(tasks)
    tasks_exhausted = False
    progress_context = (
        tqdm(
            total=len(tasks),
            desc=f"Search ({lanes} accelerator lane"
                 f"{'s' if lanes != 1 else ''})",
        )
        if show_progress
        else nullcontext(None)
    )

    with cuda_matmul_precision(precision), \
            h5py.File(input_h5, "r", libver="latest", swmr=True) as hf, \
            ThreadPoolExecutor(
                max_workers=lanes,
                thread_name_prefix="search-accelerator",
            ) as accelerator_executor, \
            ThreadPoolExecutor(
                max_workers=workers,
                thread_name_prefix="search-cpu",
            ) as cpu_executor, \
            progress_context as progress:
        while (
            not tasks_exhausted
            or accelerator_pending
            or ready_for_cpu
            or cpu_pending
        ):
            while ready_for_cpu and len(cpu_pending) < workers:
                cpu_pending.add(
                    cpu_executor.submit(
                        finish_search,
                        ready_for_cpu.popleft(),
                    )
                )

            while (
                not tasks_exhausted
                and len(accelerator_pending) < lanes
                and len(ready_for_cpu) + len(accelerator_pending) < ready_limit
            ):
                try:
                    (
                        idx,
                        header,
                        safe_h,
                        q_emb,
                        mode,
                        gap,
                        norm_mode,
                    ) = next(task_iterator)
                except StopIteration:
                    tasks_exhausted = True
                    break

                t_emb = hf["embeddings"][safe_h][:]
                accelerator_pending.add(
                    accelerator_executor.submit(
                        _compute_accelerated_search,
                        (
                            idx,
                            header,
                            q_emb,
                            t_emb,
                            mode,
                            gap,
                            norm_mode,
                            device,
                            precision,
                        ),
                    )
                )

            completed_accelerator = {
                future
                for future in accelerator_pending
                if future.done()
            }
            completed_cpu = {
                future for future in cpu_pending if future.done()
            }
            if not completed_accelerator and not completed_cpu:
                pending = accelerator_pending | cpu_pending
                if not pending:
                    continue
                completed, _ = wait(
                    pending,
                    return_when=FIRST_COMPLETED,
                )
                completed_accelerator = completed & accelerator_pending
                completed_cpu = completed & cpu_pending

            for future in completed_accelerator:
                accelerator_pending.remove(future)
                ready_for_cpu.append(future.result())
            for future in completed_cpu:
                cpu_pending.remove(future)
                results.append(future.result())
                if progress is not None:
                    progress.update(1)

    return results


def _select_lanes(tasks, workers, input_h5, device):
    manual = _configured_lanes(device, workers)
    if manual is not None:
        return manual
    candidates = _lane_candidates(device, workers)
    if len(candidates) == 1 or len(tasks) < 2:
        return 1

    cache_key = (_device_type(device), _accelerator_name(device), int(workers))
    if cache_key in accelerator_lane_cache:
        return accelerator_lane_cache[cache_key]

    count = max(2, min(int(ACCELERATOR_TUNE_PAIRS), len(tasks)))
    sample = list(tasks[:count])
    print(
        f"[Search] Auto-tuning accelerator lanes on {cache_key[1]} "
        f"using {count} representative targets..."
    )
    _run_accelerated_search(
        sample, workers, input_h5, device, 1, False
    )

    rates = {}
    for lanes in candidates:
        started = time.perf_counter()
        try:
            _run_accelerated_search(
                sample, workers, input_h5, device, lanes, False
            )
        except (RuntimeError, NotImplementedError) as error:
            if lanes == 1:
                raise
            print(f"    {lanes} lanes unavailable: {error}")
            continue
        rates[lanes] = count / max(time.perf_counter() - started, 1e-9)

    fastest = max(rates.values())
    selected = min(
        lanes
        for lanes, rate in rates.items()
        if rate >= fastest * 0.97
    )
    accelerator_lane_cache[cache_key] = selected
    print(
        "    "
        + ", ".join(
            f"{lanes}: {rate:.1f} targets/s"
            for lanes, rate in sorted(rates.items())
        )
    )
    print(f"[Search] Selected {selected} accelerator lane(s).")
    return selected


def _run_cpu_search(tasks, workers, input_h5, show_progress=True):
    results = []
    with Pool(
        processes=workers,
        initializer=init_worker,
        initargs=(input_h5,),
    ) as pool:
        iterator = pool.imap_unordered(
            search_cpu_worker,
            tasks,
            chunksize=50,
        )
        for result in tqdm(iterator, total=len(tasks), disable=not show_progress):
            results.append(result)
    return results


def _finish_fixed_query(task, query_length, target_length, matrix):
    idx, header, _safe_h, _query, mode, gap, norm_mode = task
    return finish_search(
        (
            idx,
            header,
            query_length,
            target_length,
            mode,
            gap,
            norm_mode,
            matrix,
        )
    )


def _cost_stratified_search_sample(tasks, lengths):
    count = min(256, max(16, int(len(tasks) * 0.01)))
    count = min(count, len(tasks))
    ordered = sorted(tasks, key=lambda task: (int(lengths[int(task[0])]), int(task[0])))
    if count >= len(ordered):
        return ordered
    positions = np.linspace(0, len(ordered) - 1, num=count, dtype=np.int64)
    return [ordered[int(position)] for position in positions]


def _search_results_equivalent(
    fp32_results,
    candidate_results,
    tolerance=1e-3,
    candidate_label="candidate",
):
    fp32 = {int(result["index"]): result for result in fp32_results}
    candidate_map = {
        int(result["index"]): result for result in candidate_results
    }
    if fp32.keys() != candidate_map.keys():
        return False, "target identities differ"
    for index, baseline in fp32.items():
        candidate = candidate_map[index]
        if int(baseline["aln_len"]) != int(candidate["aln_len"]):
            return False, f"alignment length changed for target {index}"
        values = (
            float(candidate["raw_score"]),
            float(candidate["norm_score"]),
        )
        baseline_values = (
            float(baseline["raw_score"]),
            float(baseline["norm_score"]),
        )
        if not all(np.isfinite(value) for value in baseline_values + values):
            return False, f"non-finite {candidate_label} result for target {index}"
        denominator = max(1, int(baseline["aln_len"]))
        drift = abs(
            float(baseline["raw_score"]) - float(candidate["raw_score"])
        ) / denominator
        if drift > tolerance:
            return False, f"per-aligned-residue drift exceeded {tolerance:g}"
    return True, "alignment lengths and per-residue scores passed"


active_search_hardware = {
    "device": "cpu",
    "plan": "scalar",
    "precision": "ieee_fp32",
    "lanes": 1,
    "microbatch_mib": None,
}


def _execute_search_plan(
    plan,
    tasks,
    workers,
    input_h5,
    store,
    lengths,
    query_embedding,
    show_progress,
):
    candidate, variant, precision, lanes = plan
    if candidate.is_cpu:
        return _run_cpu_search(tasks, workers, input_h5, show_progress)
    if variant == "tiled":
        progress = tqdm(total=len(tasks), desc="Search (tiled CUDA)") if show_progress else None
        try:
            return run_fixed_query_cuda_pipeline(
                tasks,
                query_embedding=query_embedding,
                store=store,
                lengths=lengths,
                device=candidate.device,
                workers=workers,
                lanes=lanes,
                alignment_callback=_finish_fixed_query,
                precision=precision,
                progress=progress,
            )
        finally:
            if progress is not None:
                progress.close()
    return _run_accelerated_search(
        tasks,
        workers,
        input_h5,
        candidate.device,
        lanes,
        show_progress,
        precision=precision,
    )


def _select_search_plans(tasks, workers, input_h5, store, lengths, query_embedding):
    precision_setting = normalize_precision_setting(ACCELERATOR_PRECISION)
    candidates = Hardware_Utils.get_available_devices()
    manual = Hardware_Utils.resolve_device_selection(DEVICE_SELECTION, candidates)
    if manual is not None:
        candidates = [manual]
    if precision_setting == "tf32":
        candidates = [
            candidate for candidate in candidates
            if candidate.backend == "cuda" and is_nvidia_cuda(candidate.device)
        ]
        if not candidates:
            raise RuntimeError("Forced TF32 SSEARCH requires an NVIDIA CUDA device.")
    elif precision_setting == "bf16":
        candidates = [
            candidate for candidate in candidates
            if not candidate.is_cpu
            and bf16_accelerator_support(candidate.device)[0]
        ]
        if not candidates:
            raise RuntimeError(
                "Forced BF16 SSEARCH requires a CUDA/ROCm, XPU, or MPS "
                "accelerator that passes the BF16 runtime probe."
            )

    sample = _cost_stratified_search_sample(tasks, lengths)
    print(
        f"[Hardware] Testing SSEARCH plans on {len(sample)} length-stratified "
        f"targets ({len(tasks)} total)."
    )
    print("Device/backend                 Plan      Prec.    Lanes   Targets/s   Status")
    benchmark_rows = []
    result_payloads = {}
    bf16_baselines = {}
    for candidate in candidates:
        variants = ["scalar"]
        if candidate.backend == "cuda" and len(tasks) >= TILED_SEARCH_MIN_TARGETS:
            variants.append("tiled")
        precisions = ["float32"]
        if precision_setting == "tf32":
            precisions = ["tf32"]
        elif precision_setting == "bf16":
            precisions = ["bf16"]
        elif (
            precision_setting == "automatic_32bit"
            and len(tasks) >= TF32_SEARCH_MIN_TARGETS
            and candidate.backend == "cuda"
            and is_nvidia_cuda(candidate.device)
        ):
            precisions.append("tf32")
        lane_candidates = [1] if candidate.is_cpu else _lane_candidates(candidate.device, workers)
        if precision_setting == "bf16":
            try:
                bf16_baselines[candidate.spec] = _execute_search_plan(
                    (candidate, "scalar", "float32", 1),
                    sample,
                    workers,
                    input_h5,
                    store,
                    lengths,
                    query_embedding,
                    False,
                )
            except Exception as error:
                print(
                    f"[Precision] Cannot build FP32 baseline on "
                    f"{candidate.display_name}: {error}."
                )
                Hardware_Utils.release_device_cache(candidate)
                continue
        for variant in variants:
            for precision in precisions:
                if precision == "tf32" and not is_nvidia_cuda(candidate.device):
                    continue
                for lanes in lane_candidates:
                    if variant == "tiled":
                        estimate = estimate_fixed_query_cuda_working_set(
                            sample,
                            query_embedding=query_embedding,
                            store=store,
                            lengths=lengths,
                            device=candidate.device,
                            lanes=lanes,
                            precision=precision,
                        )
                        if not estimate.feasible:
                            print(
                                f"{candidate.display_name[:30]:30}  {variant:8}  "
                                f"{precision:7}  {lanes:>5}   {'--':>9}   "
                                f"skipped: {estimate.reason}"
                            )
                            continue
                    started = time.perf_counter()
                    try:
                        payload = _execute_search_plan(
                            (candidate, variant, precision, lanes),
                            sample,
                            workers,
                            input_h5,
                            store,
                            lengths,
                            query_embedding,
                            False,
                        )
                        rate = len(sample) / max(time.perf_counter() - started, 1e-9)
                        row = Hardware_Utils.BenchmarkResult(
                            candidate, rate, lanes=lanes,
                            variant=f"{variant}:{precision}",
                        )
                        benchmark_rows.append(row)
                        result_payloads[(candidate.spec, variant, precision, lanes)] = payload
                        print(
                            f"{candidate.display_name[:30]:30}  {variant:8}  "
                            f"{precision:7}  {lanes:>5}   {rate:>9.2f}   ok"
                        )
                    except Exception as error:
                        print(
                            f"{candidate.display_name[:30]:30}  {variant:8}  "
                            f"{precision:7}  {lanes:>5}   {'--':>9}   "
                            f"{type(error).__name__}: {error}"
                        )
        Hardware_Utils.release_device_cache(candidate)

    if precision_setting == "automatic_32bit":
        fp32_rates = [
            float(row.value) for row in benchmark_rows
            if row.variant.endswith(":float32")
        ]
        tf32_rows = [row for row in benchmark_rows if row.variant.endswith(":tf32")]
        validated_tf32 = []
        for row in tf32_rows:
            variant = row.variant.split(":", 1)[0]
            baseline_rows = [
                baseline for baseline in benchmark_rows
                if baseline.candidate.spec == row.candidate.spec
                and baseline.variant == f"{variant}:float32"
            ]
            if not baseline_rows:
                continue
            baseline = max(baseline_rows, key=lambda item: float(item.value))
            equivalent, reason = _search_results_equivalent(
                result_payloads[(baseline.candidate.spec, variant, "float32", baseline.lanes)],
                result_payloads[(row.candidate.spec, variant, "tf32", row.lanes)],
            )
            if equivalent:
                validated_tf32.append(row)
            else:
                print(f"[Precision] Rejected {variant} TF32: {reason}.")
        fp32_best = max(fp32_rates, default=0.0)
        tf32_best = max((float(row.value) for row in validated_tf32), default=0.0)
        if tf32_best < fp32_best * 1.10:
            if tf32_rows:
                print(
                    f"[Precision] Using IEEE FP32: best validated TF32 speedup "
                    f"was {tf32_best / max(fp32_best, 1e-9):.2f}x."
                )
            validated_tf32 = []
        allowed_tf32 = set(id(row) for row in validated_tf32)
        benchmark_rows = [
            row for row in benchmark_rows
            if not row.variant.endswith(":tf32") or id(row) in allowed_tf32
        ]
    elif precision_setting == "bf16":
        validated_bf16 = []
        for row in benchmark_rows:
            variant, precision = row.variant.split(":", 1)
            if precision != "bf16":
                continue
            baseline = bf16_baselines.get(row.candidate.spec)
            payload = result_payloads.get(
                (row.candidate.spec, variant, precision, row.lanes)
            )
            if baseline is None or payload is None:
                continue
            equivalent, reason = _search_results_equivalent(
                baseline,
                payload,
                tolerance=BF16_PER_RESIDUE_TOLERANCE,
                candidate_label="BF16",
            )
            if equivalent:
                validated_bf16.append(row)
            else:
                print(
                    f"[Precision] Rejected {variant} BF16 on "
                    f"{row.candidate.display_name}: {reason}."
                )
        benchmark_rows = validated_bf16

    ranked_rows = Hardware_Utils.rank_benchmark_results(
        benchmark_rows, higher_is_better=True
    )
    if not ranked_rows:
        raise RuntimeError("No SSEARCH hardware plan completed successfully.")
    plans = []
    for row in ranked_rows:
        variant, precision = row.variant.split(":", 1)
        plans.append((row.candidate, variant, precision, row.lanes))
    winner = plans[0]
    print(
        f"[Hardware] Selected {winner[0].display_name}, {winner[1]} plan, "
        f"{winner[3]} lane(s), {winner[2]}."
    )
    return plans


def process_search_tasks(tasks, workers, input_h5):
    configured_precision = normalize_precision_setting(ACCELERATOR_PRECISION)
    if (
        len(tasks) < TILED_SEARCH_MIN_TARGETS
        and configured_precision != "bf16"
    ):
        device = Hardware_Utils.resolve_device_selection(
            DEVICE_SELECTION, Hardware_Utils.get_available_devices()
        )
        if device is None:
            selected_device = Hardware_Utils.get_optimal_device()
        else:
            selected_device = device.device
        precision = configured_precision
        if precision == "tf32" and not is_nvidia_cuda(selected_device):
            raise RuntimeError("Forced TF32 SSEARCH requires an NVIDIA CUDA device.")
        if _uses_accelerator(selected_device):
            lanes = _select_lanes(tasks, workers, input_h5, selected_device)
            active_search_hardware.update(
                device=str(selected_device), plan="scalar",
                precision="tf32" if precision == "tf32" else "ieee_fp32",
                lanes=lanes, microbatch_mib=None,
            )
            return _run_accelerated_search(
                tasks, workers, input_h5, selected_device, lanes, True,
                precision="tf32" if precision == "tf32" else "float32",
            )
        active_search_hardware.update(
            device="cpu", plan="scalar", precision="ieee_fp32", lanes=1,
            microbatch_mib=None,
        )
        return _run_cpu_search(tasks, workers, input_h5, True)

    store = EmbeddingTileStore(
        input_h5,
        [task[2] for task in tasks],
        host_cache_setting=0,
    )
    lengths = [shape[0] for shape in store.shapes]
    query_embedding = tasks[0][3]
    plans = _select_search_plans(
        tasks, workers, input_h5, store, lengths, query_embedding
    )
    last_error = None
    manual = Hardware_Utils.normalize_device_selection(DEVICE_SELECTION) != "auto"
    for candidate, variant, precision, lanes in plans:
        try:
            results = _execute_search_plan(
                (candidate, variant, precision, lanes),
                tasks,
                workers,
                input_h5,
                store,
                lengths,
                query_embedding,
                True,
            )
            active_search_hardware.update(
                device=candidate.display_name,
                plan=variant,
                precision=(
                    "bf16" if precision == "bf16"
                    else "tf32" if precision == "tf32"
                    else "ieee_fp32"
                ),
                lanes=lanes,
                microbatch_mib=(
                    cuda_memory_plan(
                        candidate.device,
                        lanes=lanes,
                        compute_element_bytes=precision_element_bytes(precision),
                    ).matrix_bytes
                    / (1024 ** 2)
                    if variant == "tiled" else None
                ),
            )
            return results
        except (RuntimeError, NotImplementedError, MemoryError) as error:
            last_error = error
            if manual:
                raise
            print(
                f"[Hardware] {variant} failed on {candidate.display_name}: "
                f"{error}; trying the next benchmarked plan."
            )
    raise RuntimeError("Every benchmarked SSEARCH plan failed.") from last_error

# --- 5. REPORTING -------------------------------------------------------------
METADATA_REPORT_COLUMNS = [
    "Node ID",
    "Rank",
    "Norm_Score",
    "Raw_Score",
    "Sequence_Length",
    "Alignment_Length",
]
METADATA_REPORT_TYPES = [
    "Data Type",
    "number",
    "number",
    "number",
    "number",
    "number",
]


def build_metadata_report_dataframe(rows):
    """Return search hits in the metadata viewer's two-header-row schema."""
    data_rows = [METADATA_REPORT_TYPES]
    data_rows.extend(
        [
            row["Node ID"],
            row["Rank"],
            row["Norm_Score"],
            row["Raw_Score"],
            row["Sequence_Length"],
            row["Alignment_Length"],
        ]
        for row in rows
    )
    return pd.DataFrame(data_rows, columns=METADATA_REPORT_COLUMNS)


def format_metadata_report_sheet(worksheet):
    """Match docs/metadata_template.xlsx while retaining typed cell values."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="2C3E50")
    type_fill = PatternFill("solid", fgColor="D5D8DC")
    alternate_fill = PatternFill("solid", fgColor="F7F8F9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="E2E2E2")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )
    header_font = Font(
        name="Segoe UI",
        size=11,
        bold=True,
        color="FFFFFF",
    )
    type_font = Font(
        name="Segoe UI",
        size=10,
        bold=True,
        italic=True,
        color="1F2328",
    )
    data_font = Font(name="Segoe UI", size=10, color="000000")

    worksheet.row_dimensions[1].height = 28.05
    worksheet.row_dimensions[2].height = 22.05
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = True

    widths = {
        "A": 80,
        "B": 12,
        "C": 16,
        "D": 16,
        "E": 20,
        "F": 20,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for column_index, cell in enumerate(worksheet[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = Alignment(
            horizontal="left" if column_index == 1 else "center",
            vertical="center",
        )

    for column_index, cell in enumerate(worksheet[2], start=1):
        cell.fill = type_fill
        cell.font = type_font
        cell.border = cell_border
        cell.alignment = Alignment(
            horizontal="left" if column_index == 1 else "center",
            vertical="center",
        )

    number_formats = {
        2: "0",
        3: "0.000",
        4: "0.0",
        5: "0",
        6: "0",
    }
    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=3,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=len(METADATA_REPORT_COLUMNS),
        ),
        start=3,
    ):
        worksheet.row_dimensions[row_index].height = 20
        row_fill = alternate_fill if row_index % 2 == 0 else white_fill
        for column_index, cell in enumerate(row, start=1):
            cell.fill = row_fill
            cell.font = data_font
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal="left" if column_index == 1 else "right",
                vertical="center",
            )
            if column_index in number_formats:
                cell.number_format = number_formats[column_index]


def format_search_parameters_sheet(worksheet):
    """Keep the secondary parameter sheet readable without changing its data."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="2C3E50")
    alternate_fill = PatternFill("solid", fgColor="F7F8F9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="E2E2E2")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 80
    worksheet.row_dimensions[1].height = 28.05
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            name="Segoe UI",
            size=11,
            bold=True,
            color="FFFFFF",
        )
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=2,
        ),
        start=2,
    ):
        worksheet.row_dimensions[row_index].height = 20
        row_fill = alternate_fill if row_index % 2 == 0 else white_fill
        for cell in row:
            cell.fill = row_fill
            cell.font = Font(name="Segoe UI", size=10, color="000000")
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="left", vertical="center")


def save_results(df, query_meta, db_size, seq_lookup, base_filename, query_seq, norm_mode, gap_p):
    q_head, q_len = query_meta
    
    col_header = "ALN-LEN"

    # Base metadata block
    meta_lines = []
    meta_lines.append("="*80); meta_lines.append(f"{'PROTEIN EMBEDDING SEARCH REPORT':^80}"); meta_lines.append("="*80)
    meta_lines.append(f" Query:       {q_head}")
    meta_lines.append(f" Query Len:   {q_len} residues")
    meta_lines.append(f" Database:    {INPUT_EMBED} ({db_size} sequences)")
    meta_lines.append(f" Mode:        {ALIGNMENT_MODE.upper()} Alignment")
    meta_lines.append("-" * 80)
    meta_lines.append(f" Parameters:  Gap Penalty = {gap_p}")
    meta_lines.append(f" Metric:      Raw Score / {norm_mode}")
    meta_lines.append(f" Filters:     Top_K={TOP_K} | Norm_Threshold={NORM_THRESHOLD}")
    hardware_line = (
        f" Hardware:    {active_search_hardware['device']} | "
        f"{active_search_hardware['plan']} | "
        f"{active_search_hardware['precision']} | "
        f"lanes={active_search_hardware['lanes']}"
    )
    if active_search_hardware["microbatch_mib"] is not None:
        hardware_line += (
            f" | microbatch="
            f"{active_search_hardware['microbatch_mib']:.1f} MiB"
        )
    meta_lines.append(hardware_line)
    meta_lines.append("-" * 80 + "\n")
    
    report_lines = list(meta_lines)
    onscreen_lines = list(meta_lines)
    
    if df.empty:
        report_lines.append("  [No hits found satisfying the criteria]")
        onscreen_lines.append("  [No hits found satisfying the criteria]")
        xlsx_data = []
    else:
        table_hdr = [
            f" {'RANK':<6} | {'NORM-SCR':<9} | {'RAW':<9} | {'SEQ-LEN':<8} | {col_header:<8} | {'HEADER'}",
            f"{'-'*7}-+-{'-'*9}-+-{'-'*9}-+-{'-'*8}-+-{'-'*8}-+-{'-'*35}"
        ]
        report_lines.extend(table_hdr)
        onscreen_lines.extend(table_hdr)
        
        printed_hits = 0
        rank_counter = 1
        xlsx_data = []
        for i, row in df.iterrows():
            if row['index'] == -1: continue 
            head = row['header']
            seq_len_val = int(row['seq_len'])
            aln_len_val = int(row['aln_len'])
            norm_score = row['norm_score']
            raw_score = row['raw_score']
            
            row_line = f" {rank_counter:<6} | {norm_score:<9.3f} | {raw_score:<9.1f} | {seq_len_val:<8} | {aln_len_val:<8} | {head}"
            report_lines.append(row_line)
            if printed_hits < 100:
                onscreen_lines.append(row_line)
                printed_hits += 1
                
            xlsx_data.append({
                "Node ID": str(head),
                "Rank": rank_counter,
                "Norm_Score": float(norm_score),
                "Raw_Score": float(raw_score),
                "Sequence_Length": int(seq_len_val),
                "Alignment_Length": int(aln_len_val),
            })
            
            rank_counter += 1
            
        if len(df) - 1 > 100:
            onscreen_lines.append(f"\n  [Onscreen display limited to first 100 hits. The full list of {len(df) - 1} hits is stored in the report file.]")
            
    print("\n".join(onscreen_lines))
    
    output_dir = REPORT_DIR
    os.makedirs(output_dir, exist_ok=True)
    
    report_text = "\n".join(report_lines)
    with open(os.path.join(output_dir, f"Report_{base_filename}.txt"), "w") as f: f.write(report_text)
    
    # Generate and save Excel Report
    xlsx_path = os.path.join(output_dir, f"Report_{base_filename}.xlsx")
    xlsx_df = build_metadata_report_dataframe(xlsx_data)
        
    meta_data = [
        {"Parameter": "Query Header", "Value": q_head},
        {"Parameter": "Query Length (residues)", "Value": q_len},
        {"Parameter": "Database", "Value": INPUT_EMBED},
        {"Parameter": "Database Size (sequences)", "Value": db_size},
        {"Parameter": "Alignment Mode", "Value": ALIGNMENT_MODE},
        {"Parameter": "Gap Penalty", "Value": gap_p},
        {"Parameter": "Normalization Mode", "Value": norm_mode},
        {"Parameter": "Norm Score Cutoff", "Value": NORM_THRESHOLD if NORM_THRESHOLD is not None else "None"},
        {"Parameter": "Top K", "Value": TOP_K if TOP_K is not None else "None"},
        {"Parameter": "Compute Device", "Value": active_search_hardware["device"]},
        {"Parameter": "Compute Plan", "Value": active_search_hardware["plan"]},
        {"Parameter": "Matmul Precision", "Value": active_search_hardware["precision"]},
        {"Parameter": "Accelerator Lanes", "Value": active_search_hardware["lanes"]},
        {"Parameter": "Microbatch Budget (MiB)", "Value": active_search_hardware["microbatch_mib"] if active_search_hardware["microbatch_mib"] is not None else "N/A"},
    ]
    meta_df = pd.DataFrame(meta_data)
    
    try:
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            xlsx_df.to_excel(writer, sheet_name="Search Results", index=False)
            meta_df.to_excel(writer, sheet_name="Search Parameters", index=False)
            format_metadata_report_sheet(writer.sheets["Search Results"])
            format_search_parameters_sheet(writer.sheets["Search Parameters"])
        print(f"[Export] Excel report saved to: {xlsx_path}")
    except Exception as e:
        print(f"[Warning] Failed to save Excel report: {e}")
    
    if GENERATE_FASTA:
        # OUTPUT FASTA (Ranked with query at the top)
        count = 1 
        with open(os.path.join(output_dir, f"Hits_{base_filename}.fasta"), "w") as f:
            f.write(f">{q_head}\n{query_seq}\n")
            if not df.empty and seq_lookup:
                for i, row in df.iterrows():
                    if row['index'] == -1: continue
                    head = row['header']
                    if head in seq_lookup:
                        f.write(f">{head}\n{seq_lookup[head]}\n")
                        count += 1
        print(f"[Export] {count} sequences exported to Hits_{base_filename}.fasta (Query is #1)")

# --- 6. MAIN ------------------------------------------------------------------
def main(argv=None):
    global FULL_INPUT_EMBED
    load_tool_settings(globals(), __file__, PROJECT_ROOT, argv)
    FULL_INPUT_EMBED = os.path.join(EMBED_DIR, INPUT_EMBED) if EMBED_DIR else ""
    database = prepare_database_embeddings()
    db_headers = database.headers
    seq_lookup = database.sequence_by_header

    # Process Query Input
    raw_query_name = QUERY_HEADER if QUERY_HEADER else "Manual_Query"
    query_name = sanitize_header(str(raw_query_name))[0] or "Manual_Query"
    manual_query_enabled = bool(MANUAL_QUERY_SEQ)
    query_seq_str = resolve_manual_query_sequence(
        manual_query_enabled,
        QUERY_SEQUENCE,
    )
    query_emb = None
    target_dtype = dtype_for_saving_mode(database.saving_mode)
    cleanup_device = None

    if manual_query_enabled:
        print("[Input] Using manually provided sanitized query sequence.")
    else:
        if query_name not in seq_lookup:
            raise ValueError(
                f"QUERY_SEQUENCE is empty and sanitized QUERY_HEADER "
                f"'{query_name}' is not stored in the embedding database."
            )
        query_seq_str = seq_lookup[query_name]
        with h5py.File(FULL_INPUT_EMBED, "r") as hf:
            query_emb = hf["embeddings"][query_name][:]
        print(
            f"[Input] Reusing the stored sequence and embedding for "
            f"'{query_name}'."
        )

    if query_emb is None:
        print("[Input] Generating a new embedding for the query...")
        model_obj, device, model_type = load_model_integrated(database.model_name)
        cleanup_device = device
        query_emb = get_embedding_integrated(
            query_seq_str,
            model_obj,
            device,
            model_type,
            target_dtype,
        )
        validate_embedding_array(
            query_emb,
            query_seq_str,
            database.saving_mode,
            feature_dimension=database.feature_dimension,
            require_finite=True,
            header=query_name,
        )

    # The embedding model is no longer needed. Release it before the
    # multi-lane search so smaller accelerators retain maximum working memory.
    if "model_obj" in locals():
        del model_obj
    gc.collect()
    release_accelerator_cache(cleanup_device)

    # 3. Search
    gap_p = LOCAL_GAP_P if ALIGNMENT_MODE == "local" else GLOBAL_GAP_P
    tasks = []
    
    for i, header in enumerate(db_headers):
        tasks.append((i, header, header, query_emb, ALIGNMENT_MODE, gap_p, NORM_MODE))
        
    print(
        f"[Search] Scanning {len(tasks)} sequences against "
        f"{database.model_name} using '{NORM_MODE}' normalization..."
    )
    results = []
    
    try: set_start_method('spawn')
    except RuntimeError: pass

    results = process_search_tasks(tasks, WORKERS, FULL_INPUT_EMBED)
            
    df = pd.DataFrame(results)
    df = df.sort_values(by="norm_score", ascending=False)
    
    # 4. Filter
    if NORM_THRESHOLD is not None: 
        df = df[df['norm_score'] >= float(NORM_THRESHOLD)]
        
    df = filter_ranked_hits(
        df,
        query_name,
        manual_query_enabled,
        TOP_K,
    )
    
    # Add dummy row for Query (Ensures it appears at the top of the text report)
    q_row = pd.DataFrame([{"index": -1, "header": f"(Query) {query_name}", "raw_score": 0.0, "norm_score": 99.9, "length": len(query_emb), "seq_len": len(query_emb), "aln_len": len(query_emb)}])
    df = pd.concat([q_row, df], ignore_index=True)
    
    # Use custom name if provided, otherwise fallback to sanitized query name
    if OUTPUT_NAME and str(OUTPUT_NAME).strip():
        base_filename = str(OUTPUT_NAME).strip()
    else:
        base_filename = re.sub(r'[^a-zA-Z0-9]', '_', query_name)[:20]
        
    save_results(df, (query_name, len(query_emb)), len(db_headers), seq_lookup, base_filename, query_seq_str, NORM_MODE, gap_p)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
