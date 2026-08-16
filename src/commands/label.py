# Copyright 2026 Xuebin Feng
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import Command_Engine
import os
import glob
import re
import datetime
import tempfile
import numpy as np
from collections import Counter
from dataclasses import dataclass
from types import SimpleNamespace
import matplotlib
matplotlib.use('Agg')

import matplotlib.cm as cm
import matplotlib.colors as mcolors
from Bio import AlignIO
from Bio.Align import MultipleSeqAlignment 
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
import SSN_Config as cfg
import SSN_Utils as utils
import Cache_Manifest as cache_manifest
try:
    import commands.cluster as cluster_cmd
    import commands.logo as logo_cmd
except ImportError:
    import cluster as cluster_cmd
    import logo as logo_cmd


GLOBAL_CONSERVATION_THRESHOLD = 0.97


class _FrozenSparseAlignment:
    """Detached read-only copy of the sparse alignment data needed by label."""

    def __init__(self, alignment):
        self.matrix = alignment.matrix.copy()
        self.int_to_aa = dict(alignment.int_to_aa)
        self.headers = tuple(getattr(alignment, "headers", ()))
        self.n_seqs, self.n_cols = self.matrix.shape

    def __len__(self):
        return self.n_seqs

    def __getitem__(self, index):
        row = self.matrix[index].toarray()[0]
        sequence = "".join(
            self.int_to_aa.get(int(value), "-") if value != 0 else "-"
            for value in row
        )
        description = (
            self.headers[index] if index < len(self.headers) else f"row_{index}"
        )
        return SeqRecord(
            Seq(sequence),
            id=description.split()[0],
            description=description,
        )

    def __iter__(self):
        for index in range(self.n_seqs):
            yield self[index]

    def get_alignment_length(self):
        return self.n_cols


class _FrozenAlignmentManager:
    def __init__(self, alignment, viewer_to_aln=None):
        source = alignment.aln
        if hasattr(source, "matrix"):
            self.aln = _FrozenSparseAlignment(source)
        else:
            self.aln = MultipleSeqAlignment(
                [
                    SeqRecord(
                        Seq(str(record.seq)),
                        id=record.id,
                        name=record.name,
                        description=record.description,
                    )
                    for record in source
                ]
            )
        self.col_to_label = dict(alignment.col_to_label)
        self.label_to_col = dict(alignment.label_to_col)
        self.has_reference = bool(getattr(alignment, "has_reference", False))
        self.resolved_ref_full = getattr(alignment, "resolved_ref_full", None)
        if viewer_to_aln is None:
            viewer_to_aln = getattr(alignment, "viewer_to_aln", ())
        self.viewer_to_aln = np.asarray(viewer_to_aln, dtype=int).copy()
        self.viewer_to_aln.setflags(write=False)

@dataclass(frozen=True)
class _LabelJobEnvelope:
    viewer_snapshot: object
    args: tuple


def _setting(viewer, name, default=None):
    settings = getattr(viewer, "_label_settings", None)
    if settings is not None and name in settings:
        return settings[name]
    return getattr(cfg, name, default)


def print_help():
    print("""
    Differential Labeling & Statistics Tool
    =======================================
    Generates a comprehensive XLSX report comparing the sequence properties and 
    conserved residues of each subset against the global dataset. Output is saved 
    to the 'Results/Cluster_Label/' directory.

    * PREREQUISITES: 
      1. A Multiple Sequence Alignment (MSA) must be loaded.
      2. A Reference Sequence must be set (use the 'reference' command).

    Usage: label [TARGET] [GLOBAL_MAX] [CLUSTER_MIN] [IDENTITY] [NAME]
       or: label [TARGET] [key value] [<key 2> <value 2> ...] [NAME]

    Targets (Default: clusters):
      clusters : Analyzes all defined topology clusters AND any custom groups.
      groups   : Analyzes ONLY custom groups (topology clusters not required).

    Arguments (Accepts decimals '0.4' or percentages '40%'):
      gmax (Outside Max)  : Default 40%. Max frequency a conserved residue can
                            have outside the union of all analyzed subsets where
                            that same residue meets cmin at the same position.
      cmin (Cluster Min)  : Default 98%. Min frequency a residue must have WITHIN 
                            a subset to be reported as conserved.
      id (Identity)       : Optional sequence-redundancy threshold. Equivalent
                            forms: 0.9, 90, or 90%. Reweighting is OFF unless
                            supplied. Without the 'id' keyword, identity must be
                            the third positional number after gmax and cmin.
      NAME                 : Optional final XLSX filename. '.xlsx' is added if
                            omitted. Numeric or reserved names must include the
                            extension, for example '0.4.xlsx' or 'groups.xlsx'.
                            An existing custom filename is replaced. Automatic
                            names use a numeric suffix instead of overwriting.

    Fixed behavior:
      Every amino acid meeting cmin is evaluated. Conserved subsets sharing the
      same amino acid and position use one deduplicated exclusion union; if that
      union leaves no outside sequences, the residue is not subset specific.
      Multiple passing amino acids share one workbook cell (for example,
      "Y120 | F120") in descending subset-frequency order.
      Globally conserved residues are reported when their frequency is greater
      than 97% across all aligned sequences. This threshold is not configurable.
      Label and logo jobs share one sequential background queue. Alignment,
      memberships, reference numbering, and parameters are captured on submission.

    Examples:
      label                       (Uses gmax=40%, cmin=98%, and timestamp naming)
      label 0.4 0.9               (Positional: gmax=40%, cmin=90%)
      label id 90%                 (Uses default gmax/cmin and 90% identity weights)
      label 0.4 0.9 90% report    (Sets gmax, cmin, identity, and filename)
      label groups cmin 90%       (Keyword: Analyzes groups, sets cmin to 90%)
      label groups report         (Writes report.xlsx)
      label 0.4 0.9 report        (Sets thresholds and writes report.xlsx)
      
    Note: Do not mix positional numbers after using keywords. The first two
          positional numbers remain gmax and cmin. A custom filename must be final.
    """)

def parse_percentage(val_str):
    try:
        clean_str = val_str.replace('%', '')
        val = float(clean_str)
        if val > 1.0: return val / 100.0
        return val
    except ValueError: return None


def _normalize_output_filename(filename):
    """Return a safe XLSX basename for the configured label output directory."""
    filename = str(filename).strip()
    if not filename:
        raise ValueError("Filename cannot be empty.")
    if filename in {".", ".."} or "/" in filename or "\\" in filename:
        raise ValueError("Filename must not include a directory or path separators.")
    if re.search(r'[<>:"|?*\x00-\x1f]', filename):
        raise ValueError(f"Filename contains unsupported characters: '{filename}'.")
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    return filename


def _parse_label_arguments(args):
    """Parse label arguments without reading or mutating viewer state."""
    valid_keys = {
        "gmax", "global_max", "g_max",
        "cmin", "cluster_min", "c_min",
        "id",
    }
    fixed_keys = {"gmin", "global_min", "g_min"}
    valid_targets = {"cluster", "clusters", "group", "groups"}
    forced_target = "clusters"
    requested_filename = None
    positional_args = []
    keyword_args = {}
    keyword_mode = False

    index = 0
    while index < len(args):
        raw_argument = args[index]
        argument = raw_argument.lower()
        if argument in valid_targets:
            forced_target = (
                "clusters" if argument in {"cluster", "clusters"} else "groups"
            )
            index += 1
            continue
        if argument in fixed_keys:
            raise ValueError(
                "gmin is fixed at 97% and cannot be set by the label command."
            )
        if argument in valid_keys:
            keyword_mode = True
            if index + 1 >= len(args):
                raise ValueError(f"Missing numerical value for '{argument}'.")
            if argument in {"gmax", "global_max", "g_max"}:
                key_name = "gmax"
            elif argument in {"cmin", "cluster_min", "c_min"}:
                key_name = "cmin"
            else:
                key_name = "id"
            value_text = args[index + 1]
            if key_name == "id":
                value = logo_cmd.parse_identity_threshold(value_text)
            else:
                value = parse_percentage(value_text)
                if value is None:
                    raise ValueError(
                        f"Invalid percentage '{value_text}' for '{argument}'."
                    )
            if key_name in keyword_args:
                raise ValueError(f"Duplicate assignment for '{key_name}'.")
            keyword_args[key_name] = value
            index += 2
            continue

        parsed_value = parse_percentage(argument)
        if parsed_value is not None:
            if keyword_mode:
                raise ValueError(
                    f"Ambiguous input. Positional argument '{argument}' found "
                    "after keywords."
                )
            positional_args.append((raw_argument, parsed_value))
            index += 1
            continue

        if requested_filename is not None:
            raise ValueError("Provide only one custom output filename.")
        if index != len(args) - 1:
            raise ValueError("A custom output filename must be the final argument.")
        requested_filename = _normalize_output_filename(raw_argument)
        index += 1

    positional_keys = ("gmax", "cmin", "id")
    if len(positional_args) > len(positional_keys):
        raise ValueError("Too many positional numerical arguments.")
    for position, (raw_value, parsed_value) in enumerate(positional_args):
        key_name = positional_keys[position]
        if key_name in keyword_args:
            raise ValueError(
                f"Ambiguous input. '{key_name}' defined both positionally and "
                "via keyword."
            )
        if key_name == "id":
            parsed_value = logo_cmd.parse_identity_threshold(raw_value)
        keyword_args[key_name] = parsed_value

    return {
        "global_max": keyword_args.get("gmax", 0.40),
        "cluster_min": keyword_args.get("cmin", 0.98),
        "identity_threshold": keyword_args.get("id"),
        "forced_target": forced_target,
        "requested_filename": requested_filename,
    }

def get_sequence_stats(aln, gap_chars=None):
    lengths = []
    gap_chars = set(cfg.GAP_CHARS if gap_chars is None else gap_chars)
    for record in aln:
        seq_str = str(record.seq)
        ungapped_len = sum(1 for c in seq_str if c not in gap_chars)
        lengths.append(ungapped_len)
    if not lengths: return 0, 0, 0.0, 0.0
    arr = np.array(lengths)
    return int(np.min(arr)), int(np.max(arr)), np.mean(arr), np.std(arr)


def _get_amino_acid_counts(aln, col_idx, weights=None, gap_chars=None):
    """Return non-gap amino-acid counts for one alignment column."""
    gap_chars = frozenset(cfg.GAP_CHARS if gap_chars is None else gap_chars)
    if weights is not None:
        weights = np.asarray(weights, dtype=float)
        if len(weights) != len(aln):
            raise ValueError("Sequence weights must match the alignment row count.")

        aa_counts = {}
        if hasattr(aln, 'matrix'):
            column = aln.matrix.getcol(col_idx).tocoo()
            for row_idx, aa_int in zip(column.row, column.data):
                aa = aln.int_to_aa.get(int(aa_int), 'X')
                if aa not in gap_chars:
                    aa = str(aa).upper()
                    aa_counts[aa] = aa_counts.get(aa, 0.0) + float(weights[row_idx])
        else:
            for row_idx, record in enumerate(aln):
                aa = str(record.seq[col_idx]).upper()
                if aa not in gap_chars:
                    aa_counts[aa] = aa_counts.get(aa, 0.0) + float(weights[row_idx])
        return aa_counts

    if hasattr(aln, 'matrix'):
        counts = Counter(aln.matrix[:, col_idx].data)
        aa_counts = {}
        for aa_int, count in counts.items():
            aa = aln.int_to_aa.get(aa_int, 'X')
            if aa not in gap_chars:
                aa_counts[aa] = aa_counts.get(aa, 0) + count
    else:
        raw_counts = Counter(record.seq[col_idx].upper() for record in aln)
        aa_counts = {
            aa: count
            for aa, count in raw_counts.items()
            if aa not in gap_chars
        }

    return {
        str(aa).upper(): int(count)
        for aa, count in aa_counts.items()
    }


def _get_amino_acid_frequencies(
    aln, col_idx, weights=None, gap_chars=None
):
    """Return occupancy-diluted, non-gap amino-acid frequencies for one column."""
    denominator = (
        float(np.asarray(weights, dtype=float).sum())
        if weights is not None
        else float(len(aln))
    )
    if denominator <= 0.0:
        return {}
    return {
        aa: count / denominator
        for aa, count in _get_amino_acid_counts(
            aln,
            col_idx,
            weights=weights,
            gap_chars=gap_chars,
        ).items()
    }


def _format_global_amino_acid_profile(
    aln, col_idx, frequencies=None, weights=None, gap_chars=None
):
    """Format a non-gap column profile using query.py's reporting semantics."""
    if frequencies is None:
        frequencies = _get_amino_acid_frequencies(
            aln,
            col_idx,
            weights=weights,
            gap_chars=gap_chars,
        )

    profile = []
    for aa, frequency in frequencies.items():
        percentage = frequency * 100.0
        if percentage >= 1.0:
            profile.append((aa, percentage))
    profile.sort(key=lambda item: item[1], reverse=True)

    if not profile:
        return "-"
    return " | ".join(f"{aa} {percentage:>5.1f}%" for aa, percentage in profile)


def _calculate_weighted_frequencies(aln, mapping, weights, gap_chars=None):
    """Return weighted consensus statistics and residue counts by display label."""
    weights = np.asarray(weights, dtype=float)
    if len(weights) != len(aln):
        raise ValueError("Sequence weights must match the alignment row count.")

    total_weight = float(weights.sum())
    stats = {}
    counts_by_label = {}
    if total_weight <= 0.0:
        return stats, counts_by_label

    try:
        alignment_length = aln.get_alignment_length()
    except AttributeError:
        alignment_length = aln.matrix.shape[1]

    for col_idx, label in mapping.items():
        if col_idx < 0 or col_idx >= alignment_length:
            continue
        counts = _get_amino_acid_counts(
            aln,
            col_idx,
            weights=weights,
            gap_chars=gap_chars,
        )
        counts_by_label[label] = counts
        non_gap_weight = float(sum(counts.values()))
        occupancy = non_gap_weight / total_weight
        if not counts:
            stats[label] = ('-', 0.0, 0.0)
            continue

        consensus_aa, consensus_count = max(counts.items(), key=lambda item: item[1])
        stats[label] = (
            consensus_aa,
            float(consensus_count) / total_weight,
            occupancy,
        )

    return stats, counts_by_label


def _get_indexed_amino_acid_count(
    aln,
    col_idx,
    amino_acid,
    row_indices,
    weights=None,
):
    """Return one residue's count across a deduplicated set of alignment rows."""
    row_indices = np.unique(np.asarray(row_indices, dtype=int))
    if row_indices.size == 0:
        return 0.0
    if row_indices[0] < 0 or row_indices[-1] >= len(aln):
        raise IndexError("Alignment row index is outside the available alignment.")

    if weights is None:
        selected_weights = np.ones(row_indices.size, dtype=float)
    else:
        weights = np.asarray(weights, dtype=float)
        if len(weights) != len(aln):
            raise ValueError("Sequence weights must match the alignment row count.")
        selected_weights = weights[row_indices]

    target = str(amino_acid).upper()
    if hasattr(aln, "matrix"):
        encoded = aln.matrix[row_indices].getcol(col_idx).toarray().ravel()
        matching_codes = {
            int(code)
            for code, residue in aln.int_to_aa.items()
            if str(residue).upper() == target
        }
        matches = np.fromiter(
            (int(value) in matching_codes for value in encoded),
            dtype=bool,
            count=row_indices.size,
        )
    else:
        matches = np.fromiter(
            (
                str(aln[int(row_idx)].seq[col_idx]).upper() == target
                for row_idx in row_indices
            ),
            dtype=bool,
            count=row_indices.size,
        )
    return float(selected_weights[matches].sum())


def _calculate_outside_frequency(
    amino_acid,
    global_counts,
    global_size,
    excluded_count,
    excluded_size,
):
    """Return the residue frequency after excluding a union of conserved subsets."""
    outside_size = float(global_size) - float(excluded_size)
    if outside_size <= 1e-12:
        return None

    global_count = float(global_counts.get(str(amino_acid).upper(), 0.0))
    outside_count = global_count - float(excluded_count)
    if outside_count < -1e-12:
        return None
    outside_count = max(0.0, outside_count)
    return outside_count / outside_size


def _is_subset_specific_residue(
    subset_aa,
    subset_frequency,
    subset_size,
    global_counts,
    global_size,
    cluster_min,
    global_max,
    subset_count=None,
):
    """Apply cmin and preserve the historical single-subset gmax calculation."""
    if subset_frequency < cluster_min:
        return False

    if subset_count is None:
        # Preserve the historical integer-count reconstruction when weighting is off.
        subset_count = int(round(subset_frequency * subset_size))
    else:
        subset_count = float(subset_count)
    outside_frequency = _calculate_outside_frequency(
        subset_aa,
        global_counts,
        global_size,
        subset_count,
        subset_size,
    )
    return outside_frequency is not None and outside_frequency < global_max


def _format_statistics_summary(
    network_node_count,
    aligned_node_count,
    excluded_node_count,
    effective_sequence_count=None,
):
    """Format the compact workbook statistics metadata value."""
    if effective_sequence_count is None:
        effective_sequence_count = aligned_node_count
    effective_display = f"{float(effective_sequence_count):.2f}".rstrip("0").rstrip(".")
    return (
        f"Aligned {aligned_node_count} of {network_node_count} | "
        f"Excluded {excluded_node_count} | Effective {effective_display}"
    )


def _append_workbook_metadata(
    worksheet,
    out_filename,
    ref_display,
    offset_display,
    global_list,
    network_node_count=None,
    aligned_node_count=None,
    excluded_node_count=None,
    identity_threshold=None,
    effective_sequence_count=None,
):
    worksheet.append([f"Filename: {out_filename}"])
    worksheet.append([f"Reference: {ref_display}"])
    worksheet.append([f"Alignment Offset: {offset_display}"])
    if network_node_count is not None:
        statistics_summary = _format_statistics_summary(
            network_node_count,
            aligned_node_count,
            excluded_node_count,
            effective_sequence_count,
        )
        worksheet.append([f"Statistics: {statistics_summary}"])
    if identity_threshold is not None:
        worksheet.append([f"Identity Threshold: {identity_threshold * 100:g}%"])
    worksheet.append([f"Global Conserved (>{int(GLOBAL_CONSERVATION_THRESHOLD * 100)}%)"])
    worksheet.append(global_list if global_list else ["None"])
    worksheet.append([])


def _run_label_artifact(viewer, args):
    if args and args[0].lower() == 'reset':
        Command_Engine.execute_reset(viewer, ["clusters"])
        return

    try:
        alignment = getattr(viewer, 'alignment', None)
        if alignment is None or alignment.aln is None:
            viewer.console_text.text = "Error: Global Alignment not loaded."
            print("Error: Global Alignment not loaded.")
            return

        if len(alignment.aln) == 0:
            msg = (
                "Error: The selected MSA contains no aligned rows for the current "
                "network. Label analysis is unavailable."
            )
            viewer.console_text.text = msg
            print(msg)
            return

        if not getattr(alignment, 'has_reference', False):
            viewer.console_text.text = (
                "Error: No active alignment reference. Use 'reference <ID>' with "
                "a node present in the current MSA."
            )
            print(viewer.console_text.text)
            return

        if args and args[0].lower() in ['help', '-h', '-?']:
            print_help()
            if hasattr(viewer, 'console_text'):
                viewer.console_text.text = "Help information printed to the terminal"
            return

        parameters = _parse_label_arguments(args)
        global_max = parameters["global_max"]
        cluster_min = parameters["cluster_min"]
        identity_threshold = parameters["identity_threshold"]
        forced_target = parameters["forced_target"]
        requested_filename = parameters["requested_filename"]

        # --- Validations ---
        if forced_target == "clusters" and viewer.cluster_labels is None:
            viewer.console_text.text = "Error: Run 'cluster' first."
            print("Error: Run 'cluster' first to use cluster mode.")
            return
            
        if forced_target == "groups" and getattr(viewer, 'group_labels', None) is None:
            viewer.console_text.text = "Error: No groups defined."
            print("Error: No groups defined. Use the 'group' command first.")
            return

        # --- 1. Global Statistics ---
        print("Calculating Global Stats...")
        gap_chars = frozenset(_setting(viewer, "GAP_CHARS", ("-", ".")))
        if hasattr(viewer, "_label_offset_display"):
            offset_display = viewer._label_offset_display
        else:
            offset_display = utils.get_alignment_offset_display(viewer)
        print(f"Alignment Offset: {offset_display}")
        total_global_seqs = len(viewer.alignment.aln)
        total_network_nodes = getattr(viewer, 'n_nodes', len(viewer.full_headers))
        excluded_unaligned_nodes = total_network_nodes - total_global_seqs
        g_min, g_max, g_avg, g_std = get_sequence_stats(
            viewer.alignment.aln,
            gap_chars=gap_chars,
        )

        global_weights = None
        if identity_threshold is None:
            total_global_effective_n = float(total_global_seqs)
            g_stats, _ = _calculate_weighted_frequencies(
                viewer.alignment.aln,
                viewer.alignment.col_to_label,
                np.ones(total_global_seqs, dtype=float),
                gap_chars=gap_chars,
            )
            global_count_cache = {}
        else:
            print(
                "Calculating global identity-neighbour weights at "
                f"{identity_threshold * 100:g}%..."
            )
            global_sequences = [
                str(record.seq).upper() for record in viewer.alignment.aln
            ]
            global_weights = logo_cmd.calculate_identity_weights(
                global_sequences,
                identity_threshold,
                report_backend=True,
            )
            total_global_effective_n = float(global_weights.sum())
            g_stats, global_count_cache = _calculate_weighted_frequencies(
                viewer.alignment.aln,
                viewer.alignment.col_to_label,
                global_weights,
                gap_chars=gap_chars,
            )

        global_frequency_cache = {}

        def get_global_counts(label):
            if label not in global_count_cache:
                col_idx = viewer.alignment.label_to_col.get(label)
                global_count_cache[label] = (
                    _get_amino_acid_counts(
                        viewer.alignment.aln,
                        col_idx,
                        gap_chars=gap_chars,
                    )
                    if col_idx is not None
                    else {}
                )
            return global_count_cache[label]

        def get_global_frequencies(label):
            if label not in global_frequency_cache:
                global_frequency_cache[label] = {
                    aa: count / total_global_effective_n
                    for aa, count in get_global_counts(label).items()
                } if total_global_effective_n > 0.0 else {}
            return global_frequency_cache[label]

        # --- 2. Resolve Base Directories ---
        fasta_file = _setting(viewer, 'NODE_FASTA_FILE', None)
        fasta_base = os.path.splitext(os.path.basename(fasta_file))[0] if fasta_file else _setting(viewer, 'SEQUENCE_SET', 'Network')
        metadata = getattr(viewer, "_label_network_metadata", None)
        if metadata is None:
            metadata = cache_manifest.validate_network_schema(
                _setting(viewer, 'INPUT_HDF5')
            )
        model_label = re.sub(
            r'[<>:"/\\|?*]', "_", metadata.model_name
        )
        lvl1_name = f"{fasta_base}_[{model_label}]"
        is_blast = metadata.network_type == "blast"
        if not is_blast:
            norm_m = _setting(viewer, 'NORM_MODE', None)
            if norm_m: lvl1_name += f"_{norm_m}"
            score_m = _setting(viewer, 'ALIGNMENT_SCORE', None)
            if score_m: lvl1_name += f"_{score_m}"
            
        lvl2_name_base = ""
        top_val = _setting(viewer, 'TOP_EDGE_PERCENT', None)
        if top_val is not None and str(top_val).strip() != "None":
            try: lvl2_name_base += f"Top{float(top_val)}Pct"
            except: pass
        else:
            thresh = _setting(viewer, 'SIMILARITY_THRESHOLD', 0.0)
            try: lvl2_name_base += f"Score{float(thresh)}"
            except: pass
            
        if forced_target == "clusters":
            if getattr(viewer, 'last_cluster_params', None):
                c_mode_param, c_min_param = viewer.last_cluster_params
                if lvl2_name_base:
                    lvl2_name = f"{lvl2_name_base}_{c_mode_param}_Min{c_min_param}"
                else:
                    lvl2_name = f"{c_mode_param}_Min{c_min_param}"
            else:
                lvl2_name = lvl2_name_base
        else:
            lvl2_name = lvl2_name_base

        # --- 3. Prepare Tasks ---
        tasks = [] 
        
        print(f"Splitting Global Alignment for {forced_target.upper()}...")
        viewer_to_aln, _ = Command_Engine.get_alignment_mapping(viewer)
        
        if forced_target == "clusters":
            aln_idx_to_cid = {}
            for i, _header in enumerate(viewer.full_headers):
                if i >= len(viewer.cluster_labels): break
                cid = viewer.cluster_labels[i]
                aln_idx = int(viewer_to_aln[i])
                if aln_idx >= 0:
                    aln_idx_to_cid[aln_idx] = cid

            clusters_records = {}
            for i, record in enumerate(viewer.alignment.aln):
                if i in aln_idx_to_cid:
                    found_cid = aln_idx_to_cid[i]
                    if found_cid != -1:
                        if found_cid not in clusters_records: clusters_records[found_cid] = []
                        clusters_records[found_cid].append((i, record))
            
            for cid in sorted(clusters_records.keys()):
                indexed_records = clusters_records[cid]
                sub_aln = MultipleSeqAlignment([record for _, record in indexed_records])
                aln_indices = np.asarray([idx for idx, _ in indexed_records], dtype=int)
                tasks.append((
                    'cluster', cid, sub_aln, viewer.alignment.col_to_label, aln_indices
                ))
        
        # Group splitting
        if getattr(viewer, 'group_labels', None):
            aln_idx_to_groups = {}
            for i, _header in enumerate(viewer.full_headers):
                if i >= len(viewer.group_labels): break
                groups = viewer.group_labels[i]
                aln_idx = int(viewer_to_aln[i])
                if groups and aln_idx >= 0:
                    aln_idx_to_groups[aln_idx] = groups
            
            groups_records = {}
            for i, record in enumerate(viewer.alignment.aln):
                if i in aln_idx_to_groups:
                    for g_name in aln_idx_to_groups[i]:
                        if g_name not in groups_records: groups_records[g_name] = []
                        groups_records[g_name].append((i, record))
            
            for g_name in sorted(groups_records.keys()):
                indexed_records = groups_records[g_name]
                sub_aln = MultipleSeqAlignment([record for _, record in indexed_records])
                aln_indices = np.asarray([idx for idx, _ in indexed_records], dtype=int)
                tasks.append((
                    'group', g_name, sub_aln, viewer.alignment.col_to_label, aln_indices
                ))

        # --- 4. Process Tasks ---
        master_labels = set()
        cluster_results = []
        candidate_pools = {}
        candidate_amino_acids = {}
        
        # Build color map for topology clusters using cluster_cmd
        cluster_ids = [
            entity_id
            for entity_type, entity_id, _, _, _ in tasks
            if entity_type == 'cluster'
        ]
        cluster_color_map = cluster_cmd.get_cluster_color_map(cluster_ids)

        for entity_type, entity_id, c_aln, c_map, aln_indices in tasks:
            try:
                # Calculate sequence stats
                c_size = len(c_aln)
                c_min_len, c_max_len, c_avg_len, c_std_len = get_sequence_stats(
                    c_aln,
                    gap_chars=gap_chars,
                )

                # Calculate complete residue counts and frequencies. Identity-
                # enabled subsets retain their rows' globally calculated weights.
                if global_weights is None:
                    c_weights = np.ones(c_size, dtype=float)
                else:
                    c_weights = global_weights[aln_indices]
                c_effective_n = float(c_weights.sum())
                c_stats, c_counts_by_label = _calculate_weighted_frequencies(
                    c_aln,
                    c_map,
                    c_weights,
                    gap_chars=gap_chars,
                )
                c_occ_dict = {
                    lbl: c_stats[lbl][2]
                    for lbl in utils.sort_labels(c_stats.keys())
                }
                
                # Format Output Styling
                if entity_type == 'cluster':
                    if entity_id in cluster_color_map:
                        rgb = cluster_color_map[entity_id]
                        hex_code = mcolors.to_hex(rgb)
                    else:
                        hex_code = "-"
                    name_str = f"Cluster {entity_id}"
                    sort_key = (0, entity_id)
                else:
                    hex_code = "-"
                    name_str = f"Group {entity_id}"
                    # Sort groups by count descending (-c_size), then alphabetically
                    sort_key = (1, -c_size, str(entity_id))

                result = {
                    "type": entity_type,
                    "id": entity_id,
                    "name": name_str,
                    "sort_key": sort_key,
                    "count": c_size,
                    "effective_n": c_effective_n,
                    "hex": hex_code,
                    "min": c_min_len,
                    "max": c_max_len,
                    "avg": c_avg_len,
                    "std": c_std_len,
                    "data": {},
                    "occ_data": c_occ_dict
                }
                cluster_results.append(result)

                # First pass: collect every amino acid meeting cmin. The shared
                # outside background is resolved only after every subset is known.
                if c_effective_n > 0.0:
                    for lbl in utils.sort_labels(c_counts_by_label.keys()):
                        if lbl not in g_stats:
                            continue
                        for amino_acid, count in sorted(
                            c_counts_by_label[lbl].items()
                        ):
                            frequency = float(count) / c_effective_n
                            if frequency < cluster_min:
                                continue
                            amino_acid = str(amino_acid).upper()
                            candidate_pools.setdefault(
                                (lbl, amino_acid), []
                            ).append({
                                "result": result,
                                "frequency": frequency,
                                "aln_indices": aln_indices,
                            })
                            candidate_amino_acids.setdefault(lbl, set()).add(
                                amino_acid
                            )
                    
            except Exception as e: 
                print(f"Skipping {entity_type} {entity_id} due to error: {e}")
                continue

        # Second pass: for each conserved position/residue pair, exclude the
        # deduplicated union of all qualifying cluster/group memberships.
        candidate_labels = utils.sort_labels(candidate_amino_acids.keys())
        for lbl in candidate_labels:
            amino_acids = sorted(candidate_amino_acids[lbl])
            col_idx = viewer.alignment.label_to_col.get(lbl)
            if col_idx is None:
                continue
            for amino_acid in amino_acids:
                candidates = candidate_pools[(lbl, amino_acid)]
                excluded_indices = np.unique(np.concatenate([
                    candidate["aln_indices"] for candidate in candidates
                ]))
                if global_weights is None:
                    excluded_size = float(excluded_indices.size)
                else:
                    excluded_size = float(global_weights[excluded_indices].sum())
                excluded_count = _get_indexed_amino_acid_count(
                    viewer.alignment.aln,
                    col_idx,
                    amino_acid,
                    excluded_indices,
                    weights=global_weights,
                )
                outside_frequency = _calculate_outside_frequency(
                    amino_acid,
                    get_global_counts(lbl),
                    total_global_effective_n,
                    excluded_count,
                    excluded_size,
                )
                if outside_frequency is None or outside_frequency >= global_max:
                    continue

                master_labels.add(lbl)
                for candidate in candidates:
                    candidate["result"]["data"].setdefault(lbl, []).append((
                        amino_acid,
                        candidate["frequency"],
                    ))

        # One subset-position cell may contain multiple independently passing
        # amino acids. Sort by subset frequency, then residue code for stability.
        for result in cluster_results:
            for lbl, entries in list(result["data"].items()):
                entries.sort(key=lambda item: (-item[1], item[0]))
                result["data"][lbl] = {
                    "text": " | ".join(
                        f"{amino_acid}{lbl}" for amino_acid, _frequency in entries
                    ),
                    "occ": result["occ_data"].get(lbl, 0.0),
                }

        # --- 5. Export XLSX ---
        out_path = os.path.abspath(viewer._label_output_path)
        out_dir = os.path.dirname(out_path)
        out_filename = os.path.basename(out_path)
        allow_overwrite = bool(
            getattr(viewer, "_label_allow_overwrite", False)
        )
        if not os.path.exists(out_dir): os.makedirs(out_dir)
        if not allow_overwrite and os.path.exists(out_path):
            raise FileExistsError(f"Output file already exists: {out_path}")

        global_list = []
        for lbl in utils.sort_labels(g_stats.keys()):
            aa, freq, occ = g_stats[lbl]
            if freq > GLOBAL_CONSERVATION_THRESHOLD:
                global_list.append(f"{aa}{lbl}")

        sorted_cols = utils.sort_labels(list(master_labels))
        cluster_results.sort(key=lambda x: x["sort_key"])
        all_occ_labels = utils.sort_labels(list(g_stats.keys()))
        
        ref_display = (
            getattr(viewer.alignment, 'resolved_ref_full', None)
            or getattr(viewer, 'active_reference', None)
            or "None"
        )
        
        # Prepare Metadata details
        fasta_file = _setting(viewer, 'NODE_FASTA_FILE', None)
        fasta_name = os.path.basename(fasta_file) if fasta_file else _setting(viewer, 'SEQUENCE_SET', 'N/A')
        
        network_file = _setting(viewer, 'INPUT_HDF5', None)
        network_name = os.path.basename(network_file) if network_file else "N/A"
        
        msa_file = _setting(viewer, 'MSA_FILE', None)
        alignment_name = os.path.basename(msa_file) if msa_file else "N/A"
        
        if forced_target == "clusters" and getattr(viewer, 'last_cluster_params', None):
            c_mode_param, c_min_param = viewer.last_cluster_params
            parts = c_mode_param.split('_')
            cluster_mode = parts[0] if parts else c_mode_param
            param_val = parts[1] if len(parts) > 1 else ""
            cluster_params = f"Param: {param_val}, Min Size: {c_min_param}" if param_val else f"Min Size: {c_min_param}"
        else:
            cluster_mode = "Groups" if forced_target == "groups" else "N/A"
            cluster_params = "N/A"

        label_params = (
            f"gmax_outside={int(global_max*100)}%, cmin={int(cluster_min*100)}%, "
            f"target={forced_target}"
        )
        if identity_threshold is not None:
            label_params += f", identity={identity_threshold * 100:g}%"

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            viewer.console_text.text = "Error: 'openpyxl' is required for XLSX export. Run: pip install openpyxl"
            print("Error: openpyxl not installed.")
            return

        try:
            wb = openpyxl.Workbook()
            
            # ==========================================
            # TAB 1: Meta Data
            # ==========================================
            ws_meta = wb.active
            ws_meta.title = "Meta Data"
            
            total_nodes = getattr(viewer, 'n_nodes', total_global_seqs)
            n_clusters = len([r for r in cluster_results if r['type'] == 'cluster'])
            n_groups = len([r for r in cluster_results if r['type'] == 'group'])
            if n_clusters > 0 and n_groups > 0:
                topology_str = f"{total_nodes} nodes with {n_clusters} Clusters, {n_groups} Groups"
            elif n_clusters > 0:
                topology_str = f"{total_nodes} nodes with {n_clusters} Clusters"
            else:
                topology_str = f"{total_nodes} nodes with {n_groups} Groups"

            ws_meta.append(["Metadata Field", "Value"])
            ws_meta.append(["Cluster Mode", cluster_mode])
            ws_meta.append(["Cluster Parameters", cluster_params])
            ws_meta.append(["Network Topology", topology_str])
            ws_meta.append(["Fasta Name", fasta_name])
            ws_meta.append(["Network Name", network_name])
            ws_meta.append(["Alignment Name", alignment_name])
            ws_meta.append(["Label Parameters", label_params])
            ws_meta.append([
                "Statistics",
                _format_statistics_summary(
                    total_network_nodes,
                    total_global_seqs,
                    excluded_unaligned_nodes,
                    total_global_effective_n,
                ),
            ])
            
            # Style header row for Meta Data tab
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_meta[1]:
                cell.fill = header_fill
                cell.font = header_font
            ws_meta.column_dimensions['A'].width = 25
            ws_meta.column_dimensions['B'].width = 50

            effective_percent_column = 2 if identity_threshold is not None else None
            percent_column = 3 if identity_threshold is not None else 2
            effective_n_column = 4 if identity_threshold is not None else None
            hex_color_column = 6 if identity_threshold is not None else 4
            position_start_column = 12 if identity_threshold is not None else 10
            percent_number_format = "0.00%"
            effective_n_number_format = "0.00"

            def node_fraction(count):
                return count / total_network_nodes if total_network_nodes else 0.0

            def effective_fraction(effective_n):
                if total_global_effective_n <= 0.0:
                    return 0.0
                return effective_n / total_global_effective_n

            # ==========================================
            # TAB 2: Subset Specific Matrix
            # ==========================================
            ws1 = wb.create_sheet(title="Subset Stats")
            ws1.freeze_panes = "C1"  # Keep columns A and B visible during horizontal scrolling.
            
            # Write Metadata 
            _append_workbook_metadata(
                ws1,
                out_filename,
                ref_display,
                offset_display,
                global_list,
                total_network_nodes,
                total_global_seqs,
                excluded_unaligned_nodes,
                identity_threshold,
                total_global_effective_n,
            )
            
            # Write Headers 
            ws1.append(["Subset Specific Matrix"])
            if identity_threshold is not None:
                headers1 = [
                    "Subset Name", "Effective Percent", "Percent",
                    "Effective N", "Count N",
                ]
            else:
                headers1 = ["Subset Name", "Percent", "Count"]
            headers1 += [
                "Hex Color", "Min Len", "Max Len", "Avg Len", "Std Dev", ""
            ] + [f"#{c}" for c in sorted_cols]
            ws1.append(headers1)
            
            try:
                occ_cmap1 = cm.get_cmap('Reds_r')
            except AttributeError:
                occ_cmap1 = matplotlib.colormaps['Reds_r']
                
            def get_occ_fill1(occ_value):
                rgba = occ_cmap1(occ_value)
                hex_color = mcolors.to_hex(rgba)[1:].upper()
                return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Write Global Row 
            if identity_threshold is not None:
                global_freq_row1 = [
                    "Global Stats",
                    effective_fraction(total_global_effective_n),
                    node_fraction(total_global_seqs),
                    total_global_effective_n,
                    total_global_seqs,
                ]
            else:
                global_freq_row1 = [
                    "Global Stats",
                    node_fraction(total_global_seqs),
                    total_global_seqs,
                ]
            global_freq_row1 += [
                "-", g_min, g_max, round(g_avg, 1), round(g_std, 1), ""
            ]
            
            g_occ_dict1 = {}
            for col in sorted_cols:
                if col in g_stats:
                    _, _, g_occ = g_stats[col]
                    col_idx = viewer.alignment.label_to_col[col]
                    global_freq_row1.append(
                        _format_global_amino_acid_profile(
                            viewer.alignment.aln,
                            col_idx,
                            frequencies=get_global_frequencies(col),
                            gap_chars=gap_chars,
                        )
                    )
                    g_occ_dict1[col] = g_occ
                else: 
                    global_freq_row1.append("-")
                    
            ws1.append(global_freq_row1)
            g_row_idx1 = ws1.max_row
            if effective_percent_column is not None:
                ws1.cell(
                    row=g_row_idx1, column=effective_percent_column
                ).number_format = percent_number_format
            ws1.cell(row=g_row_idx1, column=percent_column).number_format = percent_number_format
            if effective_n_column is not None:
                ws1.cell(
                    row=g_row_idx1, column=effective_n_column
                ).number_format = effective_n_number_format
            
            for c_idx, col in enumerate(sorted_cols):
                if col in g_occ_dict1:
                    col_letter_idx = c_idx + position_start_column
                    ws1.cell(row=g_row_idx1, column=col_letter_idx).fill = get_occ_fill1(g_occ_dict1[col])
                    
            ws1.append([]) # Blank row below Global Stats

            # Write Subset Rows
            last_type = None
            for res in cluster_results:
                if last_type == 'cluster' and res['type'] == 'group':
                    ws1.append([]) # Blank row separating Clusters and Groups
                last_type = res['type']
                if identity_threshold is not None:
                    row1 = [
                        res['name'],
                        effective_fraction(res['effective_n']),
                        node_fraction(res['count']),
                        res['effective_n'],
                        res['count'],
                    ]
                else:
                    row1 = [
                        res['name'],
                        node_fraction(res['count']),
                        res['count'],
                    ]
                row1 += [
                    res['hex'], res['min'], res['max'], round(res['avg'], 1),
                    round(res['std'], 1), ""
                ]
                
                row_occs1 = {}
                for c_idx, col in enumerate(sorted_cols): 
                    if col in res['data']:
                        row1.append(res['data'][col]["text"])
                    else:
                        row1.append("")
                        
                    if col in res['occ_data']:
                        row_occs1[c_idx + position_start_column] = res['occ_data'][col]
                    else:
                        row_occs1[c_idx + position_start_column] = 0.0
                        
                ws1.append(row1)
                current_row1 = ws1.max_row
                if effective_percent_column is not None:
                    ws1.cell(
                        row=current_row1, column=effective_percent_column
                    ).number_format = percent_number_format
                ws1.cell(row=current_row1, column=percent_column).number_format = percent_number_format
                if effective_n_column is not None:
                    ws1.cell(
                        row=current_row1, column=effective_n_column
                    ).number_format = effective_n_number_format
                
                if res['hex'] != "-":
                    hex_val = res['hex'].replace("#", "").upper()
                    ws1.cell(row=current_row1, column=hex_color_column).fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                
                for col_index, occ_val in row_occs1.items():
                    ws1.cell(row=current_row1, column=col_index).fill = get_occ_fill1(occ_val)
                    
            # ==========================================
            # TAB 2: Occupancy Stats
            # ==========================================
            ws2 = wb.create_sheet(title="Occupancy Stats")
            ws2.freeze_panes = "C1"  # Keep columns A and B visible during horizontal scrolling.
            
            # Write Metadata 
            _append_workbook_metadata(
                ws2,
                out_filename,
                ref_display,
                offset_display,
                global_list,
                total_network_nodes,
                total_global_seqs,
                excluded_unaligned_nodes,
                identity_threshold,
                total_global_effective_n,
            )
            
            # Write Headers 
            ws2.append(["Occupancy Matrix"])
            if identity_threshold is not None:
                headers2 = [
                    "Subset Name", "Effective Percent", "Percent",
                    "Effective N", "Count N",
                ]
            else:
                headers2 = ["Subset Name", "Percent", "Count"]
            headers2 += [
                "Hex Color", "Min Len", "Max Len", "Avg Len", "Std Dev", ""
            ] + [f"#{c}" for c in all_occ_labels]
            ws2.append(headers2)
            
            try:
                occ_cmap2 = cm.get_cmap('Greens')
            except AttributeError:
                occ_cmap2 = matplotlib.colormaps['Greens']
                
            def get_occ_fill2(occ_value):
                rgba = occ_cmap2(occ_value)
                hex_color = mcolors.to_hex(rgba)[1:].upper()
                return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Write Global Row 
            if identity_threshold is not None:
                global_freq_row2 = [
                    "Global Stats",
                    effective_fraction(total_global_effective_n),
                    node_fraction(total_global_seqs),
                    total_global_effective_n,
                    total_global_seqs,
                ]
            else:
                global_freq_row2 = [
                    "Global Stats",
                    node_fraction(total_global_seqs),
                    total_global_seqs,
                ]
            global_freq_row2 += [
                "-", g_min, g_max, round(g_avg, 1), round(g_std, 1), ""
            ]
            
            g_occ_dict2 = {}
            for col in all_occ_labels:
                global_freq_row2.append("") # Keep text blank
                if col in g_stats:
                    g_occ_dict2[col] = g_stats[col][2] # Occupancy is the 3rd item in the tuple
                else: 
                    g_occ_dict2[col] = 0.0
                    
            ws2.append(global_freq_row2)
            g_row_idx2 = ws2.max_row
            if effective_percent_column is not None:
                ws2.cell(
                    row=g_row_idx2, column=effective_percent_column
                ).number_format = percent_number_format
            ws2.cell(row=g_row_idx2, column=percent_column).number_format = percent_number_format
            if effective_n_column is not None:
                ws2.cell(
                    row=g_row_idx2, column=effective_n_column
                ).number_format = effective_n_number_format
            
            for c_idx, col in enumerate(all_occ_labels):
                col_letter_idx = c_idx + position_start_column
                ws2.cell(row=g_row_idx2, column=col_letter_idx).fill = get_occ_fill2(g_occ_dict2[col])
                
            ws2.append([]) # Blank row below Global Stats

            # Write Subset Rows
            last_type = None
            for res in cluster_results:
                if last_type == 'cluster' and res['type'] == 'group':
                    ws2.append([]) # Blank row separating Clusters and Groups
                last_type = res['type']
                if identity_threshold is not None:
                    row2 = [
                        res['name'],
                        effective_fraction(res['effective_n']),
                        node_fraction(res['count']),
                        res['effective_n'],
                        res['count'],
                    ]
                else:
                    row2 = [
                        res['name'],
                        node_fraction(res['count']),
                        res['count'],
                    ]
                row2 += [
                    res['hex'], res['min'], res['max'], round(res['avg'], 1),
                    round(res['std'], 1), ""
                ]
                
                row_occs2 = {}
                for c_idx, col in enumerate(all_occ_labels): 
                    row2.append("") # Keep text blank
                    row_occs2[c_idx + position_start_column] = res['occ_data'].get(col, 0.0)
                        
                ws2.append(row2)
                current_row2 = ws2.max_row
                if effective_percent_column is not None:
                    ws2.cell(
                        row=current_row2, column=effective_percent_column
                    ).number_format = percent_number_format
                ws2.cell(row=current_row2, column=percent_column).number_format = percent_number_format
                if effective_n_column is not None:
                    ws2.cell(
                        row=current_row2, column=effective_n_column
                    ).number_format = effective_n_number_format
                
                if res['hex'] != "-":
                    hex_val = res['hex'].replace("#", "").upper()
                    ws2.cell(row=current_row2, column=hex_color_column).fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                
                for col_index, occ_val in row_occs2.items():
                    ws2.cell(row=current_row2, column=col_index).fill = get_occ_fill2(occ_val)

            # Auto-fit Column A width based on the longest cluster or group name
            name_lengths = [len("Subset Name"), len("Global Stats")] + [len(res['name']) for res in cluster_results]
            max_name_len = max(name_lengths) if name_lengths else 15
            col_a_width = max(max_name_len + 3, 15)

            ws1.column_dimensions['A'].width = col_a_width
            ws2.column_dimensions['A'].width = col_a_width
            ws1.column_dimensions['B'].width = 18 if identity_threshold is not None else 10
            ws2.column_dimensions['B'].width = 18 if identity_threshold is not None else 10
            if identity_threshold is not None:
                ws1.column_dimensions['C'].width = 10
                ws2.column_dimensions['C'].width = 10

            file_descriptor, partial_path = tempfile.mkstemp(
                prefix=f".{os.path.splitext(out_filename)[0]}.",
                suffix=".partial.xlsx",
                dir=out_dir,
            )
            os.close(file_descriptor)
            try:
                wb.save(partial_path)
                if not allow_overwrite and os.path.exists(out_path):
                    raise FileExistsError(
                        f"Output file already exists: {out_path}"
                    )
                os.replace(partial_path, out_path)
                partial_path = None
            finally:
                if partial_path and os.path.exists(partial_path):
                    try:
                        os.remove(partial_path)
                    except OSError:
                        pass
            
            msg = f"Exported to {out_path}"
            viewer.console_text.text = msg
            print(msg)
            return {
                "message": msg,
                "save_path": out_path,
                "reveal_directory": out_dir,
            }
        except Exception as e:
            viewer.console_text.text = f"IO Error: {e}"
            raise

    except Exception as e:
        viewer.console_text.text = f"Error: {e}"
        raise


def _available_automatic_output(scheduler, directory, filename):
    stem, suffix = os.path.splitext(filename)
    candidate = filename
    index = 2
    while True:
        path = os.path.abspath(os.path.join(directory, candidate))
        if not os.path.exists(path) and not scheduler.is_output_path_reserved(path):
            return candidate, path
        candidate = f"{stem}_{index}{suffix}"
        index += 1


def _execute_label_envelope(envelope):
    result = _run_label_artifact(
        envelope.viewer_snapshot,
        list(envelope.args),
    )
    if not isinstance(result, dict):
        message = getattr(envelope.viewer_snapshot.console_text, "text", "")
        raise RuntimeError(message or "Label generation did not produce an artifact.")
    return result


def _report_label_error(viewer, error):
    message = f"Error: {error}"
    if hasattr(viewer, "console_text"):
        viewer.console_text.text = message
    if hasattr(viewer, "update_console_background"):
        viewer.update_console_background()
    print(message)


def run(viewer, args):
    """Validate and snapshot label inputs before enqueuing the heavy work."""
    if args and args[0].lower() == "reset":
        Command_Engine.execute_reset(viewer, ["clusters"])
        return

    alignment = getattr(viewer, "alignment", None)
    if alignment is None or alignment.aln is None:
        _report_label_error(viewer, "Global Alignment not loaded.")
        return
    if len(alignment.aln) == 0:
        _report_label_error(
            viewer,
            "The selected MSA contains no aligned rows for the current network. "
            "Label analysis is unavailable.",
        )
        return
    if not getattr(alignment, "has_reference", False):
        _report_label_error(
            viewer,
            "No active alignment reference. Use 'reference <ID>' with a node "
            "present in the current MSA.",
        )
        return
    if args and args[0].lower() in {"help", "-h", "-?"}:
        print_help()
        if hasattr(viewer, "console_text"):
            viewer.console_text.text = "Help information printed to the terminal"
        return

    try:
        parameters = _parse_label_arguments(args)
    except ValueError as error:
        _report_label_error(viewer, error)
        return

    forced_target = parameters["forced_target"]
    if forced_target == "clusters" and getattr(viewer, "cluster_labels", None) is None:
        _report_label_error(viewer, "Run 'cluster' first.")
        return
    if forced_target == "groups" and getattr(viewer, "group_labels", None) is None:
        _report_label_error(viewer, "No groups defined.")
        return

    scheduler = getattr(viewer, "background_job_scheduler", None)
    if scheduler is None:
        _report_label_error(viewer, "The background job scheduler is unavailable.")
        return

    output_directory = os.path.abspath(
        getattr(cfg, "CLUSTER_LABEL_DIR", os.path.join("Results", "Cluster_Label"))
    )
    requested_filename = parameters["requested_filename"]
    allow_overwrite = requested_filename is not None
    if requested_filename is None:
        generated = (
            "Label_Output_"
            + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            + ".xlsx"
        )
        output_filename, output_path = _available_automatic_output(
            scheduler,
            output_directory,
            generated,
        )
    else:
        output_filename = requested_filename
        output_path = os.path.abspath(
            os.path.join(output_directory, output_filename)
        )
        if scheduler.is_output_path_reserved(output_path):
            _report_label_error(
                viewer,
                f"Output file is already reserved by a background job: {output_path}",
            )
            return

    try:
        network_metadata = cache_manifest.validate_network_schema(cfg.INPUT_HDF5)
        viewer_to_aln, _ = Command_Engine.get_alignment_mapping(viewer)
        frozen_alignment = _FrozenAlignmentManager(alignment, viewer_to_aln)
    except Exception as error:
        _report_label_error(viewer, f"Could not snapshot label inputs: {error}")
        return

    group_labels = getattr(viewer, "group_labels", None)
    if group_labels is not None:
        group_labels = tuple(
            frozenset(groups) if groups else frozenset()
            for groups in group_labels
        )
    cluster_labels = getattr(viewer, "cluster_labels", None)
    if cluster_labels is not None:
        cluster_labels = tuple(cluster_labels)

    settings = {
        name: getattr(cfg, name, default)
        for name, default in (
            ("NODE_FASTA_FILE", None),
            ("SEQUENCE_SET", "Network"),
            ("INPUT_HDF5", None),
            ("MSA_FILE", None),
            ("NORM_MODE", None),
            ("ALIGNMENT_SCORE", None),
            ("TOP_EDGE_PERCENT", None),
            ("SIMILARITY_THRESHOLD", 0.0),
            ("GAP_CHARS", ("-", ".")),
        )
    }
    snapshot = SimpleNamespace(
        alignment=frozen_alignment,
        active_reference=getattr(viewer, "active_reference", None),
        alignment_offset=getattr(viewer, "alignment_offset", 0),
        full_headers=tuple(getattr(viewer, "full_headers", ())),
        n_nodes=int(getattr(viewer, "n_nodes", len(getattr(viewer, "full_headers", ())))),
        cluster_labels=cluster_labels,
        group_labels=group_labels,
        last_cluster_params=(
            tuple(viewer.last_cluster_params)
            if getattr(viewer, "last_cluster_params", None)
            else None
        ),
        console_text=SimpleNamespace(text=""),
        _label_settings=settings,
        _label_network_metadata=network_metadata,
        _label_offset_display=utils.get_alignment_offset_display(viewer),
        _label_output_path=output_path,
        _label_allow_overwrite=allow_overwrite,
    )
    envelope = _LabelJobEnvelope(snapshot, tuple(args))
    try:
        scheduler.enqueue(
            command_name="label",
            description=f"label -> {output_filename}",
            payload=envelope,
            worker=_execute_label_envelope,
            output_path=output_path,
            allow_overwrite=allow_overwrite,
        )
    except (FileExistsError, RuntimeError) as error:
        _report_label_error(viewer, error)
